from django.shortcuts import get_object_or_404, render
from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAuthenticated
from exams.tasks import *
from usermgmt.models import UserProfile
from .serializers import *
from . models import *
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from usermgmt.custompagination import CustomPagination
from .permissions import *
from rest_framework import permissions, status
from branches.models import *
from rest_framework.exceptions import NotFound
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError, NotFound, ParseError
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from usermgmt.authentication import QueryParameterTokenAuthentication
from rest_framework.authentication import SessionAuthentication
from django.http import StreamingHttpResponse, HttpResponse
import csv
import io
from progresscard.models import ExamProgressCardMapping

# ---------------- Subject ----------------
class SubjectDropdownViewSet(ModelViewSet):
    queryset = Subject.objects.filter(is_active=True).order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = SubjectDropdownSerializer
    http_method_names = ['get']

# ---------------- Subject Category Dropdown----------------
class SubjectCategoryDropdownViewSet(ModelViewSet):
    queryset = SubjectCategory.objects.filter(is_active=True).order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = SubjectCategoryDropdownSerializer
    http_method_names = ['get']

# ---------------- Subject Category Dropdown----------------
class ExamCategoryDropdownViewSet(ModelViewSet):
    queryset = ExamCategory.objects.filter(is_active=True).order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamCategoryDropdownSerializer
    http_method_names = ['get']

# class SubjectDropdownForExamInstanceViewSet(ModelViewSet):
#     # queryset = Subject.objects.filter(is_active=True).order_by('name')
#     permission_classes = [IsAuthenticated]
#     serializer_class = SubjectDropdownSerializer
#     http_method_names = ['get']

#     def get_queryset(self):
#         exam_id = self.kwargs.get('exam_id')
#         if not exam_id:
#             return Response({'exam_id': "This field is required in the URL."}, status=status.HTTP_400_BAD_REQUEST)
        
#         classes = Exam.objects.filter(exam_id=exam_id).values_list('student_classes', flat=True)
#         subjects =Subject.objects.filter(class_names__in=classes).distinct()

#         return subjects


class SubjectDropdownForExamInstanceViewSet(ModelViewSet):
    """
    Provides a dropdown list of subjects associated with the classes of a given Exam.
    Includes the subject of the current ExamInstance if `exam_instance_id` is provided (for update view).
    URL pattern: /subject_dropdown_for_exam_instance/<exam_id>/?exam_instance_id=<id>&category_id=<id>
    """
    permission_classes = [IsAuthenticated]
    serializer_class = SubjectDropdownSerializer
    http_method_names = ['get']

    def get_queryset(self):
        exam_id = self.kwargs.get('exam_id')
        category_id = self.request.query_params.get('category_id')
        exam_instance_id = self.request.query_params.get('exam_instance_id')

        if not exam_id:
            raise ValidationError({'exam_id': "This field is required in the URL."})

        exam = (
            Exam.objects.filter(exam_id=exam_id, is_active=True)
            .prefetch_related('student_classes')
            .first()
        )
        if not exam:
            raise ValidationError({'exam_id': f"Exam with ID {exam_id} not found or inactive."})

        class_names = exam.student_classes.all()
        if not class_names.exists():
            return Subject.objects.none()

        # Optional category filter
        category = None
        if category_id:
            try:
                category = SubjectCategory.objects.get(id=category_id)
            except SubjectCategory.DoesNotExist:
                raise ValidationError({'category_id': 'Category not found.'})

        # Subjects already assigned to this exam (except current instance if editing)
        exam_subjects_qs = ExamInstance.objects.filter(exam=exam, is_active=True)
        if exam_instance_id:
            exam_subjects_qs = exam_subjects_qs.exclude(exam_instance_id=exam_instance_id)
        exam_subjects = exam_subjects_qs.values_list('subject__subject_id', flat=True)

        # Subjects available for all classes of the exam
        subjects_qs = (
            Subject.objects.filter(is_active=True, class_names__in=class_names)
            .annotate(class_count=Count('class_names', filter=Q(class_names__in=class_names), distinct=True))
            .filter(class_count=class_names.count())
            .exclude(subject_id__in=exam_subjects)
            .distinct()
            .order_by('name')
        )

        # Apply category filter if provided
        if category:
            subjects_qs = subjects_qs.filter(category_id=category.id)

        return subjects_qs


# class SubjectDropdownForExamInstanceViewSet(ModelViewSet):
#     """
#     Provides a dropdown list of subjects associated with the classes of a given Exam.
#     Includes the subject of the current ExamInstance if `exam_instance_id` is provided (for update view).
#     URL pattern: /subject_dropdown_for_exam_instance/<exam_id>/?exam_instance_id=<id>
#     """
#     permission_classes = [IsAuthenticated]
#     serializer_class = SubjectDropdownSerializer
#     http_method_names = ['get']
 
#     def get_queryset(self):
#         exam_id = self.kwargs.get('exam_id')
#         exam_instance_id = self.request.query_params.get('exam_instance_id')  # ✅ optional for update
#         if not exam_id:
#             raise ValidationError({'exam_id': "This field is required in the URL."})
 
#         exam = (Exam.objects.filter(exam_id=exam_id, is_active=True).prefetch_related('student_classes').first())
 
#         if not exam:
#             raise ValidationError({'exam_id': f"Exam with ID {exam_id} not found or inactive."})
 
#         class_names = exam.student_classes.all()
 
#         if not class_names.exists():
#             return Subject.objects.none()
       
#         if exam_instance_id:
#             exam_subjects = ExamInstance.objects.filter(exam=exam, is_active=True).values_list('subject__subject_id', flat=True).exclude(exam_instance_id=exam_instance_id)
#         else:
#         # Subjects already assigned in active ExamInstances
#             exam_subjects =  ExamInstance.objects.filter(exam=exam, is_active=True).values_list('subject__subject_id', flat=True)
 
#         # ✅ Base queryset: subjects matching all classes
#         subjects_qs = Subject.objects.filter(
#             is_active=True,
#             class_names__in=class_names
#         ).annotate(
#             class_count=Count('class_names', filter=Q(class_names__in=class_names), distinct=True)
#         ).filter(
#             class_count=class_names.count()
#         ).exclude(subject_id__in=exam_subjects).distinct().order_by('name')
 
#         return subjects_qs


# ---------------- SubjectSkill ----------------
class SubjectSkillDropdownViewSet(ModelViewSet):
    queryset = SubjectSkill.objects.filter(is_active=True).order_by('subject')
    permission_classes = [IsAuthenticated]
    serializer_class = SubjectSkillDropdownSerializer
    http_method_names = ['get']
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['subject']

class SubjectSkillDropdownForExamInstanceViewSet(ModelViewSet):
    queryset = SubjectSkill.objects.filter(is_active=True).order_by('subject')
    permission_classes = [IsAuthenticated]
    serializer_class = SubjectSkillDropdownSerializer
    http_method_names = ['get']
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['subject']


# ---------------- ExamType ----------------
class ExamTypeDropdownViewSet(ModelViewSet):
    queryset = ExamType.objects.filter(is_active=True).order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamTypeDropdownSerializer
    http_method_names = ['get']
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = []

# ---------------- Exam ----------------
class ExamDropdownViewSet(ModelViewSet):
    queryset = Exam.objects.filter(is_active=True).order_by('-start_date')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamDropdownSerializer
    http_method_names = ['get']


# ---------------- ExamInstance ----------------
class ExamInstanceDropdownViewSet(ModelViewSet):
    queryset = ExamInstance.objects.filter(is_active=True).order_by('-date')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamInstanceDropdownSerializer
    http_method_names = ['get']


# ---------------- ExamAttendanceStatus ----------------
class ExamAttendanceStatusDropdownViewSet(ModelViewSet):
    queryset = ExamAttendanceStatus.objects.filter(is_active=True).order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamAttendanceStatusDropdownSerializer
    http_method_names = ['get']

# ---------------- ExamResultStatus ----------------
class ExamResultStatusDropdownViewSet(ModelViewSet):
    queryset = ExamResultStatus.objects.all().order_by('name')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamResultStatusDropdownSerializer
    http_method_names = ['get']

# ==================== Subject ====================
class SubjectViewSet(ModelViewSet):
    queryset = Subject.objects.filter(is_active=True).order_by('name')
    serializer_class = SubjectSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'display_name', 'description', 'category__name']  # text fields to search
    filterset_fields = ['academic_devisions', 'class_names', 'is_active', 'category']  # FK/many2many and boolean
    ordering_fields = ['name', 'display_name', 'created_at', 'updated_at', 'category__name']  # fields users can order by
    pagination_class = CustomPagination

    def perform_create(self, serializer):
            serializer.save(created_by=self.request.user, updated_by=self.request.user)
    def perform_update(self, serializer):
            serializer.save(updated_by=self.request.user)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewSubject]
        elif self.action == 'create':
            permission_classes = [CanAddSubject]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeSubject]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]


# ==================== SubjectSkill ====================
class SubjectSkillViewSet(ModelViewSet):
    queryset = SubjectSkill.objects.filter(is_active=True).order_by('subject')
    serializer_class = SubjectSkillSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'subject__name']  # searchable text fields
    filterset_fields = ['subject', 'is_active']  # FK and boolean fields
    ordering_fields = ['name', 'subject__name', 'created_at', 'updated_at']  # sortable fields
    pagination_class = CustomPagination

    def perform_create(self, serializer):
            serializer.save(created_by=self.request.user, updated_by=self.request.user)
    def perform_update(self, serializer):
            serializer.save(updated_by=self.request.user)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewSubjectSkill]
        elif self.action == 'create':
            permission_classes = [CanAddSubjectSkill]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeSubjectSkill]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]


# ==================== ExamType ====================
class ExamTypeViewSet(ModelViewSet):
    queryset = ExamType.objects.filter(is_active=True).order_by('name')
    serializer_class = ExamTypeSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'description']       # text search
    filterset_fields = ['is_active']             # boolean filter
    ordering_fields = ['name', 'created_at', 'updated_at']  # sortable fields
    pagination_class = CustomPagination

    def perform_create(self, serializer):
            serializer.save(created_by=self.request.user, updated_by=self.request.user)
    def perform_update(self, serializer):
            serializer.save(updated_by=self.request.user)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewExamType]
        elif self.action == 'create':
            permission_classes = [CanAddExamType]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeExamType]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

# ==================== Grade Boundary ====================
class GradeBoundaryViewSet(ModelViewSet):
    # queryset = GradeBoundary.objects.filter(is_active=True).order_by('-min_percentage')
    serializer_class = GradeBoundarySerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['grade', 'remarks', 'category__name']
    filterset_fields = ['category', 'is_active']
    ordering_fields = ['min_percentage', 'max_percentage', 'grade', 'category__name', 'remarks']
    pagination_class = CustomPagination

    def get_queryset(self):
        category_id = self.request.query_params.get('category_id')
        if not category_id:
            raise ParseError(
                {"detail": "The 'category_id' query parameter is required."}
            )
        return GradeBoundary.objects.filter(is_active=True, category_id=category_id).order_by('-min_percentage')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewGradeBoundary]
        elif self.action == 'create':
            permission_classes = [CanAddGradeBoundary]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeGradeBoundary]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

# ==================== Co-Scholastic Grade ====================
class CoScholasticGradeViewSet(ModelViewSet):
    # queryset = CoScholasticGrade.objects.filter(is_active=True).order_by('-point')
    serializer_class = CoScholasticGradeSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'description', 'category__name']
    filterset_fields = ['category', 'is_active']
    ordering_fields = ['name', 'point', 'created_at', 'updated_at', 'category__name']
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def get_queryset(self):
        category_id = self.request.query_params.get('category_id')
        if not category_id:
            raise ParseError(
                {"detail": "The 'category_id' query parameter is required."}
            )
        return CoScholasticGrade.objects.filter(is_active=True, category_id=category_id).order_by('name')

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewCoScholasticGrade]
        elif self.action == 'create':
            permission_classes = [CanAddCoScholasticGrade]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeCoScholasticGrade]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

#
# ==================== Exam ====================
class ExamViewSet(ModelViewSet):
    serializer_class = ExamSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['name', 'exam_type__name', 'academic_year__name', 'exam_status__name', 'category__name']
    filterset_fields = [
        'exam_type', 'is_visible', 'is_progress_card_visible',
        'is_active', 'academic_year', 'name', 'start_date',
        'end_date', 'marks_entry_expiry_datetime', 'exam_status', 'category',
    ]
    ordering_fields = [
        'exam_type__name', 'start_date', 'end_date', 'name',
        'is_visible', 'created_at', 'updated_at', 'exam_status__name',
        'academic_year', 'marks_entry_expiry_datetime', 'category__name',
    ]
    pagination_class = CustomPagination

    def get_queryset(self):
        current_academic_year = AcademicYear.objects.filter(is_current_academic_year=True).first()
        if not current_academic_year:
            raise NotFound("Current academic year not found.")
        return (
            Exam.objects.filter(academic_year=current_academic_year, is_active=True)
            .order_by('-exam_id')
        )

    def perform_create(self, serializer):
        """
        Wrap creation in a DB transaction and catch duplicate constraint errors
        """
        try:
            with transaction.atomic():
                serializer.save(created_by=self.request.user, updated_by=self.request.user)
        except IntegrityError as e:
            # Handle unique constraint violation gracefully
            if 'unique_exam_per_year_type' in str(e):
                raise serializers.ValidationError({
                    "name": "An exam with this name, academic year, and exam type already exists."
                })
            # Handle academic_year auto-selection issues
            if 'academic_year_id' in str(e):
                raise serializers.ValidationError({
                    "academic_year": "Valid academic year not found or inactive."
                })
            raise serializers.ValidationError({
                "non_field_errors": "A database integrity error occurred."
            })

    def perform_update(self, serializer):
        """
        Prevent IntegrityError on updates too
        """
        try:
            with transaction.atomic():
                serializer.save(updated_by=self.request.user)
        except IntegrityError as e:
            if 'unique_exam_per_year_type' in str(e):
                raise serializers.ValidationError({
                    "name": "An exam with this name, academic year, and exam type already exists."
                })
            raise serializers.ValidationError({
                "non_field_errors": "A database integrity error occurred during update."
            })

    def create(self, request, *args, **kwargs):
        """
        Override create to return consistent 400 instead of 500
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            self.perform_create(serializer)
        except serializers.ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """
        Override update for safe IntegrityError handling
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        try:
            self.perform_update(serializer)
        except serializers.ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # def perform_create(self, serializer):
    #     validated_data = serializer.validated_data
    #     name = validated_data.get("name")
    #     exam_type = validated_data.get("exam_type")

    #     # === Validation checks ===
    #     if not name or not str(name).strip():
    #         raise ValidationError({"name": "Exam name is required."})

    #     if not exam_type:
    #         raise ValidationError({"exam_type": "Exam Type is required."})

    #     current_academic_year = AcademicYear.objects.filter(is_current_academic_year=True).first()
    #     if not current_academic_year:
    #         raise ValidationError({"academic_year": "Current academic year not found."})

    #     # === Prevent duplicates ===
    #     if Exam.objects.filter(
    #         name__iexact=name.strip(),
    #         academic_year=current_academic_year,
    #         exam_type=exam_type,
    #     ).exists():
    #         raise ValidationError({"name": "An exam with this name, year, and type already exists."})

    #     # === Safe transaction save ===
    #     try:
    #         with transaction.atomic():
    #             serializer.save(
    #                 academic_year=current_academic_year,
    #                 created_by=self.request.user,
    #                 updated_by=self.request.user
    #             )
    #     except IntegrityError:
    #         # This catches DB constraint violations gracefully
    #         raise ValidationError({"detail": "Duplicate exam entry or integrity constraint violated."})

    # def perform_update(self, serializer):
    #     try:
    #         with transaction.atomic():
    #             serializer.save(updated_by=self.request.user)
    #     except IntegrityError:
    #         raise ValidationError({"detail": "Exam update failed due to a database constraint violation."})

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewExam]
        elif self.action == 'create':
            permission_classes = [CanAddExam]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeExam]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]


# ==================== ExamInstance ====================
class ExamInstanceViewSet(ModelViewSet):
    serializer_class = ExamInstanceSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter,OrderingFilter]
    search_fields = [
        'subject__name',
        'exam__name',
        'exam__exam_type__name',
        'subject_category__name',
    ]
    filterset_fields = [
        'subject',
        'has_external_marks',
        'has_internal_marks',
        'has_subject_skills',
        'has_subject_co_scholastic_grade',
        'subject_category',
    ]
    ordering_fields = [
        'date',
        'exam_start_time',
        'exam_end_time',
        'subject__name',
        'has_external_marks',
        'has_internal_marks',
        'has_subject_skills',
        'subject_category',
    ]
    pagination_class = CustomPagination
    lookup_field = 'pk'
    lookup_url_kwarg = 'pk'

    # ✅ Get exam_id from URL kwargs
    def get_exam_id(self):
        exam_id = self.kwargs.get('exam_id')
        if not exam_id:
            raise ValidationError({"exam_id": "This field is required in the URL."})
        return exam_id

    def get_queryset(self):
        exam_id = self.get_exam_id()
        return ExamInstance.objects.filter(
            exam__exam_id=exam_id,
            is_active=True
        ).order_by('sequence')

    # # ✅ Override list() to include overall exam info
    # def list(self, request, *args, **kwargs):
    #     queryset = self.filter_queryset(self.get_queryset())
    #     page = self.paginate_queryset(queryset)

    #     serializer = self.get_serializer(page, many=True)

    #     # Get exam name once (all instances have same exam)
    #     exam_name = None
    #     if queryset.exists():
    #         exam_name = queryset.first().exam.name

    #     paginated_data = self.get_paginated_response(serializer.data).data

    #     # Inject exam_name into the overall response
    #     paginated_data['exam_name'] = exam_name

    #     return Response(paginated_data, status=status.HTTP_200_OK)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewExamInstance]
        elif self.action == 'create':
            permission_classes = [CanAddExamInstance]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeExamInstance]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

    # ✅ No need to revalidate serializer here
    # def perform_create(self, serializer):
    #     exam_id = self.get_exam_id()
    #     exam = get_object_or_404(Exam, pk=exam_id)

    #     if not exam.is_editable:
    #         raise ValidationError({"non_field_errors": "This Exam is already published — creation/edit not allowed."})

    #     subject = serializer.validated_data.get("subject")

    #     # Check for duplicates at application level first
    #     if ExamInstance.objects.filter(subject=subject, exam=exam).exists():
    #         raise ValidationError({"subject": "An exam for this subject already exists for this exam."})

    #     try:
    #         with transaction.atomic():
    #             serializer.save(exam=exam, created_by=self.request.user, updated_by=self.request.user)
    #     except IntegrityError:
    #         # Catch DB-level constraint violation and raise as 400
    #         raise ValidationError({"non_field_errors": "An exam for this subject already exists (database constraint)."})

    # def perform_update(self, serializer):
    #     exam_id = self.get_exam_id()
    #     exam = get_object_or_404(Exam, pk=exam_id)
    #     if not exam.is_editable:
    #         raise ValidationError({"non_field_errors": "This Exam is already published — creation/edit not allowed."})
    #     serializer.save(exam=exam, updated_by=self.request.user)


# class ExamInstanceViewSet(ModelViewSet):
#     serializer_class = ExamInstanceSerializer
#     http_method_names = ['get', 'post', 'put']
#     filter_backends = [DjangoFilterBackend, SearchFilter]
#     search_fields = ['exam__name', 'subject__name']
#     pagination_class = CustomPagination

#     def get_queryset(self):
#         # Only require exam_id for GET (list) requests
#         if self.action == 'list':
#             exam_id = self.request.query_params.get('exam_id')
#             if not exam_id:
#                 raise NotFound("Exam ID is required for listing exam instances.")
            
#             return ExamInstance.objects.filter(
#                 exam__exam_id=exam_id,
#                 is_active=True
#             ).order_by('date')
        
#         # For other actions (retrieve, create, update)
#         return ExamInstance.objects.filter(is_active=True).order_by('date')
    
#     def perform_create(self, serializer):
#         if serializer.is_valid():
#             serializer.save(created_by=self.request.user, updated_by=self.request.user)
#     def perform_update(self, serializer):
#         if serializer.is_valid():
#             serializer.save(updated_by=self.request.user)

#     def get_permissions(self):
#         if self.action in ['list', 'retrieve']:
#             permission_classes = [CanViewExamInstance]
#         elif self.action == 'create':
#             permission_classes = [CanAddExamInstance]
#         elif self.action in ['update', 'partial_update']:
#             permission_classes = [CanChangeExamInstance]
#         else:
#             permission_classes = [permissions.AllowAny]
#         return [permission() for permission in permission_classes]

# ==================== ExamSubjectSkillInstance ====================
class ExamSubjectSkillInstanceViewSet(ModelViewSet):
    serializer_class = ExamSubjectSkillInstanceSerializer
    http_method_names = ['get', 'post', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter,OrderingFilter]

    search_fields = [
        'subject_skill__name',              # search by skill name
        'subject_skill__subject__name',     # search by related subject name
    ]

    filterset_fields = [
        'subject_skill',
        'has_external_marks',
        'has_internal_marks',
        'has_subject_co_scholastic_grade',
        'subject_skill__name',
        'subject_skill__subject__name',
    ]

    ordering_fields = [
        'subject_skill__subject__name',
        'subject_skill__name',
        'has_external_marks',
        'has_internal_marks',
        'has_subject_co_scholastic_grade',
    ]

    pagination_class = CustomPagination
    lookup_field = 'pk'
    lookup_url_kwarg = 'pk'

    def get_exam_instance_id(self):
        """
        Get exam_instance_id from URL kwargs
        """
        exam_instance_id = self.kwargs.get('exam_instance_id')
        if not exam_instance_id:
            raise ValidationError({"exam_instance_id": "This field is required in the URL."})
        return exam_instance_id

    def get_queryset(self):
        exam_instance_id = self.get_exam_instance_id()
        exam_instance = ExamInstance.objects.get(exam_instance_id=exam_instance_id)
        return ExamSubjectSkillInstance.objects.filter(
            exam_instance=exam_instance,  # ✅ direct FK reference
            is_active=True
        ).order_by('subject_skill__subject__name')

    def perform_update(self, serializer):
        exam_instance_id = self.get_exam_instance_id()
        exam_instance = ExamInstance.objects.get(exam_instance_id=exam_instance_id)

        serializer.save(
            exam_instance=exam_instance,
            updated_by=self.request.user
        )

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewExamSubjectSkillInstance]
        elif self.action == 'create':
            permission_classes = [CanAddExamSubjectSkillInstance]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeExamSubjectSkillInstance]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

# class ExamSubjectSkillInstanceViewSet(ModelViewSet):
#     serializer_class = ExamSubjectSkillInstanceSerializer
#     http_method_names = ['get', 'put']
#     filter_backends = [DjangoFilterBackend, SearchFilter]
#     search_fields = [
#         'exam_instance__exam__name',
#         'subject_skill__subject__name',
#         'subject_skill__skill__name',
#     ]
#     filterset_fields = [
#         'exam_instance',
#         'subject_skill',
#         'has_external_marks',
#         'has_internal_marks',
#         'has_subject_co_scholastic_grade',
#         'is_active',
#     ]
#     pagination_class = CustomPagination

#     def get_queryset(self):
#         """
#         Filter ExamSubjectSkillInstance based on exam_instance_id for list views.
#         """
#         if self.action == 'list':
#             exam_instance_id = self.request.query_params.get('exam_instance_id')
#             if not exam_instance_id:
#                 raise NotFound("Exam Instance ID is required for listing exam subject skill instances.")

#             return ExamSubjectSkillInstance.objects.filter(
#                 exam_instance__exam_instance_id=exam_instance_id,
#                 is_active=True
#             ).order_by('subject_skill__subject__name')

#         # For other actions (retrieve, create, update)
#         return ExamSubjectSkillInstance.objects.filter(is_active=True).order_by('subject_skill__subject__name')
    
#     def perform_create(self, serializer):
#         if serializer.is_valid():
#             serializer.save(created_by=self.request.user, updated_by=self.request.user)
#     def perform_update(self, serializer):
#         if serializer.is_valid():
#             serializer.save(updated_by=self.request.user)

#     def get_permissions(self):
#         if self.action in ['list', 'retrieve']:
#             permission_classes = [CanViewExamSubjectSkillInstance]
#         elif self.action == 'create':
#             permission_classes = [CanAddExamSubjectSkillInstance]
#         elif self.action in ['update', 'partial_update']:
#             permission_classes = [CanChangeExamSubjectSkillInstance]
#         else:
#             permission_classes = [permissions.AllowAny]
#         return [permission() for permission in permission_classes]


class BranchWiseExamResultStatusViewSet(ModelViewSet):
    serializer_class = BranchWiseExamResultStatusSerializer
    http_method_names = ['get', 'put']
    filter_backends = [DjangoFilterBackend, SearchFilter,OrderingFilter]
    pagination_class = CustomPagination
    search_fields = [
        'academic_year__name',
        'branch__name',
        'exam__name',
        'status__name',
        'exam__exam_type__name',
        'is_progress_card_downloaded',
        'exam__category__name',
    ]

    filterset_fields = [
        'academic_year',
        'branch',
        'exam',
        'status',
        'is_visible',
        'is_active',
        'exam__exam_type',
        'is_progress_card_downloaded',
        'exam__category',

    ]

    ordering_fields = [
        'academic_year__name',
        'branch__name',
        'exam__name',
        'status__name',
        'marks_entry_expiry_datetime',
        'marks_completion_percentage',
        'updated_at',
        'exam__exam_type__name',
        'is_progress_card_downloaded',
        'exam__category__name',
    ]

    def get_queryset(self):
        user = self.request.user

        # ✅ Efficient branching logic
        if user.groups.filter(id=1).exists():  # Super admin or system user
            branches = Branch.objects.filter(is_active=True)
        else:
            branches = (
                UserProfile.objects.filter(user=user)
                .values_list('branches', flat=True)
                .distinct()
            )

        current_academic_year = AcademicYear.objects.filter(is_current_academic_year=True).first()
        if not current_academic_year:
            raise NotFound("Current academic year not found.")

        # ✅ Avoid returning inactive or invalid records
        queryset = (
            BranchWiseExamResultStatus.objects.filter(
                is_active=True,
                branch__in=branches,
                academic_year=current_academic_year,
            )
            .select_related('academic_year', 'branch', 'exam', 'status')  # optimization
            .order_by('-academic_year', 'marks_completion_percentage')
        )
        return queryset

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewBranchWiseExamResultStatus]
        elif self.action == 'create':
            permission_classes = [CanAddBranchWiseExamResultStatus]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeBranchWiseExamResultStatus]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

class SectionWiseExamResultStatusViewSet(ModelViewSet):
    """
    Handles viewing and updating Section-wise Exam Result Status entries.
    """
    serializer_class = SectionWiseExamResultStatusSerializer
    http_method_names = ['get']
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    pagination_class = CustomPagination

    search_fields = [
        'academic_year__name',
        'section__name',
        'section__class_name__name',     
        'section__orientation__name',     
        'status__name',
    ]

    filterset_fields = [
        'academic_year',
        'section__name',
        'section__class_name',         
        'section__orientation',           
        'status',
        'is_visible',
        'is_active',
        'is_progress_card_downloaded',
        'marks_entry_expiry_datetime'
    ]

    ordering_fields = [
        'academic_year__name',
        'section__name',
        'section__class_name__name',     
        'section__orientation__name',     
        'status__name',
        'marks_entry_expiry_datetime',
        'marks_completion_percentage',
        'updated_at',
        'is_progress_card_downloaded',
    ]


    def get_queryset(self):
        """
        Filters queryset by branch_id and exam_id passed in the URL.
        """
        branch_id = self.kwargs.get('branch_id')
        exam_id = self.kwargs.get('exam_id')

        if not branch_id:
            return Response({'branch_id': "This field is required in the URL."}, status=status.HTTP_400_BAD_REQUEST)
        if not exam_id:
            return Response({'exam_id': "This field is required in the URL."}, status=status.HTTP_400_BAD_REQUEST)
        
        current_academic_year = AcademicYear.objects.filter(is_current_academic_year=True).first()
        if not current_academic_year:
            raise NotFound("Current academic year not found.")
        
        queryset = (
            SectionWiseExamResultStatus.objects.filter(
                is_active=True,
                branch__branch_id=branch_id,
                exam__exam_id=exam_id,
                academic_year=current_academic_year,
            )
            .select_related(
                'academic_year',
                'branch',
                'section',
                'exam',
                'status',
            )
            .order_by('section__class_name__class_sequence', 'section__name')  # show most recent first
        )

        return queryset

    def get_permissions(self):
        """
        Assigns permission classes based on the current action.
        """
        if self.action in ['list', 'retrieve']:
            permission_classes = [CanViewSectionWiseExamResultStatus]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [CanChangeSectionWiseExamResultStatus]
        else:
            permission_classes = [permissions.AllowAny]
        return [permission() for permission in permission_classes]

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def update_section_wise_exam_result_status_view(request):
    """
    Syncs section-wise exam result status from branch-wise exam result status
    for a given branch and exam.
    - Updates existing records.
    - Creates missing SectionWiseExamResultStatus records if not found.
    """
    branch_wise_exam_result_status_id = request.query_params.get('branch_wise_exam_result_status_id')
    if not branch_wise_exam_result_status_id:
        return Response({'branch_wise_exam_result_status_id': "This field is required in the URL."}, status=status.HTTP_400_BAD_REQUEST)

    # ✅ Get branch-wise record
    branch_status = BranchWiseExamResultStatus.objects.select_related('branch', 'exam', 'academic_year').filter(
        id=branch_wise_exam_result_status_id,
        is_active=True
    ).first()
    if not branch_status:
        return Response({'branch_wise_exam_result_status_id': "Invalid Branch Wise Exam Result Status ID."}, status=status.HTTP_400_BAD_REQUEST)
   
    if not branch_status.academic_year:
        return Response({'academic_year': "Academic year is missing for this record."}, status=status.HTTP_400_BAD_REQUEST)

    # ✅ Get all matching sections
    sections = Section.objects.filter(
        academic_year=branch_status.academic_year,
        branch=branch_status.branch,
        class_name__class_name_id__in=branch_status.exam.student_classes.values_list('class_name_id', flat=True),
        orientation__orientation_id__in=branch_status.exam.orientations.values_list('orientation_id', flat=True),
        is_active=True,
        has_students=True,
    ).distinct()

    if not sections.exists():
        return Response({
            "message": ["No sections found matching the given branch, exam classes, and orientations."]
        }, status=status.HTTP_404_NOT_FOUND)

    # ✅ Bulk update existing records
    existing_records = SectionWiseExamResultStatus.objects.filter(
        section__in=sections,
        branch=branch_status.branch,
        exam=branch_status.exam,
        is_active=True,
        academic_year=branch_status.academic_year
    )

    updated_count = existing_records.update(
        # is_progress_card_downloaded=branch_status.is_progress_card_downloaded,
        marks_entry_expiry_datetime=branch_status.marks_entry_expiry_datetime,
        is_visible=branch_status.is_visible,
        updated_at=timezone.now()
    )

    # ✅ Identify and create missing records
    existing_section_ids = existing_records.values_list('section__section_id', flat=True)
    missing_sections = sections.exclude(section_id__in=existing_section_ids)

    created_count = 0
    if missing_sections.exists():
        new_objects = [
            SectionWiseExamResultStatus(
                academic_year=branch_status.academic_year,
                branch=branch_status.branch,
                section=section,
                exam=branch_status.exam,
                marks_entry_expiry_datetime=branch_status.marks_entry_expiry_datetime,
                is_visible=branch_status.is_visible,
                # is_progress_card_downloaded=branch_status.is_progress_card_downloaded,
                is_active=True,
                status = ExamResultStatus.objects.get(id=1),
            )
            for section in missing_sections
        ]
        SectionWiseExamResultStatus.objects.bulk_create(new_objects)
        created_count = len(new_objects)

    # ✅ Response
    if updated_count == 0 and created_count == 0:
        msg = "All section-wise exam result statuses are already up to date."
    else:
        msg = f"{updated_count} section-wise records updated, {created_count} created successfully."

    return Response({
        "success": True,
        "message": msg,
        "branch": branch_status.branch.name,
        "exam": branch_status.exam.name,
        "academic_year": branch_status.academic_year.name
    }, status=status.HTTP_200_OK)



#====================================================================================================================================================================
#=========================================================        EXAMS OPERATIONS        ===========================================================================
#====================================================================================================================================================================

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db import transaction
from .models import Exam, BranchWiseExamResultStatus
from exams.models import ExamResultStatus
from exams.utils.exam_visibility import *


class ExamMakeVisibleAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def put(self, request, pk):
        try:
            exam = Exam.objects.get(exam_id = pk)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found"}, status=404)

        result = set_exam_visibility(exam, user=request.user, visible=True)
        status_code = 200 if result["success"] else 400
        return Response(result, status=status_code)



class ExamMakeInvisibleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            exam = Exam.objects.get(exam_id = pk)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found"}, status=404)

        result = set_exam_visibility(exam, user=request.user, visible=False)
        return Response(result, status=200)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Exam
from django.utils import timezone

class PublishExamAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        exam_id = request.query_params.get("exam_id")

        if not exam_id:
            return Response({"detail": "exam_id is required."},status=status.HTTP_400_BAD_REQUEST,)
        try:
            exam = Exam.objects.get(exam_id = exam_id)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found."},status=status.HTTP_404_NOT_FOUND,)
    
        if exam.marks_entry_expiry_datetime and exam.marks_entry_expiry_datetime < timezone.now():
            return Response({"message": "Cannot publish. Marks entry expiry datetime has already passed.",},status=status.HTTP_400_BAD_REQUEST,)
         
        try:
            exam_status = ExamStatus.objects.get(id =2)
        except ExamStatus.DoesNotExist:
            return Response({"detail": "ExamStatus with id=2 not found."},status=status.HTTP_400_BAD_REQUEST,)
        
        # === Validate Exam Instances ===
        exam_instances = ExamInstance.objects.filter(exam=exam, is_active=True)
        if not exam_instances.exists():
            return Response(
                {'detail': 'There are no subjects for this exam.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # === Check if any instance date is outside exam date range ===
        invalid_instances = exam_instances.filter(
            Q(date__lt=exam.start_date) | Q(date__gt=exam.end_date)
        )

        if invalid_instances.exists():
            first_invalid = invalid_instances.first()
            return Response(
                {
                    'detail': f"The exam date for {first_invalid.subject.name} ({first_invalid.date}) "
                    f"is not within the scheduled exam period."
                },status=status.HTTP_400_BAD_REQUEST
            )
        
        exam.is_visible = True
        exam.exam_status = exam_status
        exam.is_editable  = False
        exam.updated_by = request.user
        exam.save(update_fields=["is_visible", "exam_status","is_editable", "updated_by"])

        branch_updated_count = 0

        with transaction.atomic():
            for branch in exam.branches.all():
                obj, created = BranchWiseExamResultStatus.objects.get_or_create(
                    academic_year =exam.academic_year,
                    branch=branch,
                    exam=exam,
                    defaults={
                        "is_visible": True,
                        "marks_entry_expiry_datetime": exam.marks_entry_expiry_datetime,
                    },
                )
                if not created:
                    obj.is_visible = True
                    obj.marks_entry_expiry_datetime = exam.marks_entry_expiry_datetime
                    obj.save(update_fields=["is_visible", "marks_entry_expiry_datetime"])
                else:
                    branch_updated_count += 1

        return Response({"message": f"Exam '{exam.name}' published successfully.","branches_processed": branch_updated_count,},status=status.HTTP_200_OK,)




class ExpireExamAPIView(APIView):
   
    permission_classes = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        exam_id = request.query_params.get("exam_id")
        if not exam_id:
            return Response({"detail": "exam_id is required."},status=status.HTTP_400_BAD_REQUEST,)
        try:
            exam = Exam.objects.get(exam_id=exam_id)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found."},status=status.HTTP_404_NOT_FOUND,)

        now = timezone.now()

        if not exam.marks_entry_expiry_datetime:
            return Response({"message": "This exam has no expiry datetime set."},status=status.HTTP_400_BAD_REQUEST,)

        if exam.marks_entry_expiry_datetime > now:
            return Response({"message": "Exam is still active. Expiry time not reached yet."},status=status.HTTP_400_BAD_REQUEST,)
        
        try:
            locked_status = ExamStatus.objects.get(id = 3)
        except ExamStatus.DoesNotExist:
            return Response({"detail": "ExamStatus 'Marks Entry Locked' not found. Please create it."},status=status.HTTP_400_BAD_REQUEST,)

        exam.is_visible = False
        exam.exam_status = locked_status
        exam.is_editable  = False
        exam.updated_by = request.user
        exam.save(update_fields=["is_visible", "exam_status", "is_editable","updated_by"])
       

        branch_updated_count = 0
        with transaction.atomic():
            branches = BranchWiseExamResultStatus.objects.filter(exam=exam, is_visible=True)
            for branch in branches:
                branch.is_visible = False
                branch.marks_entry_expiry_datetime = exam.marks_entry_expiry_datetime
            BranchWiseExamResultStatus.objects.bulk_update(branches, ["is_visible", "marks_entry_expiry_datetime"])
            branch_updated_count = branches.count()
            # for branch in BranchWiseExamResultStatus.objects.filter(exam=exam, is_visible=True):
            #     branch.is_visible = False
            #     branch.marks_entry_expiry_datetime = exam.marks_entry_expiry_datetime
            #     branch.save(update_fields=["is_visible", "marks_entry_expiry_datetime"])
            #     branch_updated_count += 1

        return Response(
            {
                "message": f"Exam '{exam.name}' expired successfully.",
                "branches_updated": branch_updated_count,
            },
            status=status.HTTP_200_OK,
        )

# from django.utils import timezone
# from django.db import transaction
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from rest_framework.permissions import IsAuthenticated
# from .models import Exam, BranchWiseExamResultStatus


class PublishProgressCardAPIView(APIView):
 
    permission_classes = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        exam_id = request.query_params.get("exam_id")

        # ✅ Validate exam_id
        if not exam_id:
            return Response({"detail": "exam_id is required."},status=status.HTTP_400_BAD_REQUEST,)

        # ✅ Get exam
        try:
            exam = Exam.objects.get(exam_id=exam_id)
        except Exam.DoesNotExist:
            return Response({"detail": "Exam not found."},status=status.HTTP_404_NOT_FOUND,)

        # ✅ Check if exam is expired
        # now = timezone.now()
        # if exam.marks_entry_expiry_datetime and exam.marks_entry_expiry_datetime < now:
        #     return Response(
        #         {"message": "Cannot publish progress card. Exam has expired."},
        #         status=status.HTTP_400_BAD_REQUEST,
        #     )

        # ✅ Update exam visibility fields
        exam.is_progress_card_visible = True
        exam.updated_by = request.user
        exam.is_editable  = False
        exam.save(update_fields=["is_progress_card_visible","is_editable", "updated_by"])

        # ✅ Optionally, you can also mark all BranchWiseExamResultStatus as progress card visible
        # (only if you maintain a similar field there)
        branch_updated_count = 0
        with transaction.atomic():
            branches = BranchWiseExamResultStatus.objects.filter(exam=exam)
            for branch in branches:
                branch.is_progress_card_downloaded = True  # 👈 optional — use your field name
            BranchWiseExamResultStatus.objects.bulk_update(
                branches, ["is_progress_card_downloaded"]
            )
            branch_updated_count = branches.count()

        return Response(
            {
                "message": f"Progress card published for exam '{exam.name}'.",
                "branches_updated": branch_updated_count,
            },
            status=status.HTTP_200_OK,
        )
#=============================================================================================================
#============================================ Marks Entry Page ===============================================
#=============================================================================================================
# @api_view(['GET'])
# @permission_classes([CanAddExamResult])
# def create_exam_results(request):
#     section_status_id = request.query_params.get('section_wise_exam_result_status_id')
#     if not section_status_id:
#         return Response({'section_wise_exam_result_status_id': "This field is required in the URL."},
#                         status=status.HTTP_400_BAD_REQUEST)

#     try:
#         section_status = SectionWiseExamResultStatus.objects.select_related('exam', 'section').get(
#             id=section_status_id, is_active=True
#         )
#     except SectionWiseExamResultStatus.DoesNotExist:
#         return Response({'section_wise_exam_result_status_id': "Invalid id"},
#                         status=status.HTTP_400_BAD_REQUEST)

#     exam = section_status.exam
#     exam_instances = ExamInstance.objects.filter(exam=exam, is_active=True) 
#     students = Student.objects.filter(
#         section=section_status.section,
#         is_active=True,
#         academic_year=exam.academic_year,
#     ).exclude(admission_status__admission_status_id=3,)

#     if not students:
#         exam_result_status = ExamResultStatus.objects.get(id=4)
#         section_status.marks_completion_percentage = 100
#         section_status.status = exam_result_status
#         section_status.save(update_fields=["marks_completion_percentage", "status"])
        
#         return Response(
#             {"Students": "No Students Found"},
#             status=status.HTTP_200_OK
#         )

#     existing_results = ExamResult.objects.filter(
#         student__in=students,
#         exam_instance__in=exam_instances,
#         is_active=True
#     ).select_related('student', 'exam_instance', 'exam_attendance', 'co_scholastic_grade')

#     results_dict = {(res.student_id, res.exam_instance_id): res for res in existing_results}

#     # Create missing ExamResults
#     new_results = []
#     for instance in exam_instances:
#         for student in students:
#             key = (student.student_id, instance.exam_instance_id)
#             if key not in results_dict:
#                 new_result = ExamResult.objects.create(student=student, exam_instance=instance)
#                 results_dict[key] = new_result
#                 new_results.append(new_result)

#     # Handle skill results
#     for instance in exam_instances.filter(has_subject_skills=True):
#         skills = instance.subject_skills.all()
#         for student in students:
#             res = results_dict.get((student.student_id, instance.exam_instance_id))
#             for skill in skills:
#                 ExamSkillResult.objects.get_or_create(exam_result=res, skill=skill)

#     # Build response
#     data = []
#     for student in students:
#         student_dict = {
#             'student_id': student.student_id,
#             'student_name': student.name,
#             'SCS_Number': student.SCS_Number,
#             'exam_instances': []
#         }

#         # Build response
#         for instance in exam_instances:
#             res = results_dict.get((student.student_id, instance.exam_instance_id))
#             skills_data = []
#             if instance.has_subject_skills:
#                 for skill in instance.subject_skills.all():
#                     skill_instance = ExamSubjectSkillInstance.objects.filter(
#                         exam_instance=instance, subject_skill=skill, is_active=True
#                     ).first()

#                     skill_result = ExamSkillResult.objects.filter(
#                         exam_result=res, skill=skill
#                     ).first()

#                     skills_data.append({
#                         'skill_name': skill.name,
#                         'has_external_marks': skill_instance.has_external_marks if skill_instance else False,
#                         'has_internal_marks': skill_instance.has_internal_marks if skill_instance else False,
#                         'has_subject_co_scholastic_grade': skill_instance.has_subject_co_scholastic_grade if skill_instance else False,
#                         'exam_skill_result_id': skill_result.exam_skill_result_id if skill_result else None,
#                         'exam_attendance': skill_result.exam_attendance.exam_attendance_status_id if skill_result and skill_result.exam_attendance else None,
#                         'max_cut_off_marks_external' : skill_instance.cut_off_marks_external if skill_instance else 0,
#                         'external_marks': (
#                             skill_result.external_marks
#                             if skill_result and skill_result.exam_attendance and skill_result.exam_attendance.exam_attendance_status_id == 1
#                             else (skill_result.exam_attendance.short_code if skill_result and skill_result.exam_attendance else None)
#                         ),
#                         'max_cut_off_marks_internal': skill_instance.cut_off_marks_internal if skill_instance else 0,
#                         'internal_marks': skill_result.internal_marks if skill_result else None,
#                         'co_scholastic_grade': skill_result.co_scholastic_grade.id if skill_result and skill_result.co_scholastic_grade else None,
#                     })

#             student_dict['exam_instances'].append({
#                 'subject_name': instance.subject.name,
#                 'exam_result_id': res.exam_result_id if res else None,
#                 'has_external_marks': instance.has_external_marks,
#                 'has_internal_marks': instance.has_internal_marks,
#                 'has_subject_co_scholastic_grade': instance.has_subject_co_scholastic_grade,
#                 'exam_attendance': res.exam_attendance.exam_attendance_status_id if res and res.exam_attendance else None,
#                 'max_cut_off_marks_external': instance.cut_off_marks_external,
#                 'external_marks': (
#                     res.external_marks
#                     if res and res.exam_attendance and res.exam_attendance.exam_attendance_status_id == 1
#                     else (res.exam_attendance.short_code if res and res.exam_attendance else None)
#                 ),
#                 'max_cut_off_marks_internal': instance.cut_off_marks_internal,
#                 'internal_marks': res.internal_marks if res else None,
#                 'co_scholastic_grade': res.co_scholastic_grade.id if res and res.co_scholastic_grade else None,
#                 'has_subject_skills': instance.has_subject_skills,
#                 'exam_skill_instances': skills_data,
#             })

#         data.append(student_dict)

#     return Response(data)

@api_view(['GET'])
@permission_classes([CanAddExamResult])
def create_exam_results(request):
    section_status_id = request.query_params.get('section_wise_exam_result_status_id')
    search = request.query_params.get('search')  # <-- New parameter for search

    if not section_status_id:
        return Response({'section_wise_exam_result_status_id': "This field is required in the URL."},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        section_status = SectionWiseExamResultStatus.objects.select_related('exam', 'section').get(
            id=section_status_id, is_active=True
        )
    except SectionWiseExamResultStatus.DoesNotExist:
        return Response({'section_wise_exam_result_status_id': "Invalid id"},
                        status=status.HTTP_400_BAD_REQUEST)

    exam = section_status.exam
    exam_instances = ExamInstance.objects.filter(exam=exam, is_active=True).order_by('sequence')

    students = Student.objects.filter(
        section=section_status.section,
        is_active=True,
        academic_year=exam.academic_year,
    ).exclude(admission_status__admission_status_id=3)

    # 🔍 Apply search if provided
    if search:
        students = students.filter(
            Q(name__icontains=search) | Q(SCS_Number__icontains=search)
        )

    if not students.exists():
        exam_result_status = ExamResultStatus.objects.get(id=4)
        section_status.marks_completion_percentage = 100
        section_status.status = exam_result_status
        section_status.save(update_fields=["marks_completion_percentage", "status"])
        # return Response({"Students": "No Students Found"}, status=status.HTTP_200_OK)
        return Response([])

    existing_results = ExamResult.objects.filter(
        student__in=students,
        exam_instance__in=exam_instances,
        is_active=True
    ).select_related('student', 'exam_instance', 'exam_attendance', 'co_scholastic_grade')

    results_dict = {(res.student_id, res.exam_instance_id): res for res in existing_results}

    # Create missing ExamResults
    new_results = []
    for instance in exam_instances:
        for student in students:
            key = (student.student_id, instance.exam_instance_id)
            if key not in results_dict:
                new_result = ExamResult.objects.create(student=student, exam_instance=instance)
                results_dict[key] = new_result
                new_results.append(new_result)

    # Handle skill results
    for instance in exam_instances.filter(has_subject_skills=True):
        skills = instance.subject_skills.all()
        for student in students:
            res = results_dict.get((student.student_id, instance.exam_instance_id))
            for skill in skills:
                ExamSkillResult.objects.get_or_create(exam_result=res, skill=skill)

    # Build response
    data = []
    for student in students:
        student_dict = {
            'student_id': student.student_id,
            'student_name': student.name,
            'SCS_Number': student.SCS_Number,
            'exam_instances': []
        }

        for instance in exam_instances:
            res = results_dict.get((student.student_id, instance.exam_instance_id))
            skills_data = []
            if instance.has_subject_skills:
                for skill in instance.subject_skills.all():
                    skill_instance = ExamSubjectSkillInstance.objects.filter(
                        exam_instance=instance, subject_skill=skill, is_active=True
                    ).first()

                    skill_result = ExamSkillResult.objects.filter(
                        exam_result=res, skill=skill
                    ).first()

                    skills_data.append({
                        'skill_name': skill.name,
                        'has_external_marks': skill_instance.has_external_marks if skill_instance else False,
                        'has_internal_marks': skill_instance.has_internal_marks if skill_instance else False,
                        'has_subject_co_scholastic_grade': skill_instance.has_subject_co_scholastic_grade if skill_instance else False,
                        'exam_skill_result_id': skill_result.exam_skill_result_id if skill_result else None,
                        'exam_attendance': skill_result.exam_attendance.exam_attendance_status_id if skill_result and skill_result.exam_attendance else None,
                        'max_cut_off_marks_external': skill_instance.cut_off_marks_external if skill_instance else 0,
                        'external_marks': (
                            skill_result.external_marks
                            if skill_result and skill_result.exam_attendance and skill_result.exam_attendance.exam_attendance_status_id == 1
                            else (skill_result.exam_attendance.short_code if skill_result and skill_result.exam_attendance else None)
                        ),
                        'max_cut_off_marks_internal': skill_instance.cut_off_marks_internal if skill_instance else 0,
                        'internal_marks': skill_result.internal_marks if skill_result else None,
                        'co_scholastic_grade': skill_result.co_scholastic_grade.id if skill_result and skill_result.co_scholastic_grade else None,
                    })

            student_dict['exam_instances'].append({
                'subject_name': instance.subject.name,
                'exam_result_id': res.exam_result_id if res else None,
                'has_external_marks': instance.has_external_marks,
                'has_internal_marks': instance.has_internal_marks,
                'has_subject_co_scholastic_grade': instance.has_subject_co_scholastic_grade,
                'exam_attendance': res.exam_attendance.exam_attendance_status_id if res and res.exam_attendance else None,
                'max_cut_off_marks_external': instance.cut_off_marks_external,
                'external_marks': (
                    res.external_marks
                    if res and res.exam_attendance and res.exam_attendance.exam_attendance_status_id == 1
                    else (res.exam_attendance.short_code if res and res.exam_attendance else None)
                ),
                'max_cut_off_marks_internal': instance.cut_off_marks_internal,
                'internal_marks': res.internal_marks if res else None,
                'co_scholastic_grade': res.co_scholastic_grade.id if res and res.co_scholastic_grade else None,
                'has_subject_skills': instance.has_subject_skills,
                'exam_skill_instances': skills_data,
            })

        data.append(student_dict)

    return Response(data)


# @api_view(['PUT'])
# @permission_classes([CanChangeExamResult])
# def edit_exam_results(request, exam_result_id):
#     """
#     Update an ExamResult instance (full update only).
#     Handles external/internal marks, co-scholastic grade, 
#     and automatically sets attendance based on marks.
#     """
#     try:
#         exam_result = ExamResult.objects.get(exam_result_id=exam_result_id, is_active=True)
#     except ExamResult.DoesNotExist:
#         return Response(
#             {"exam_result_id": "Invalid Exam Result ID"},
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # Only full update is allowed now
#     serializer = EditExamResultSerializer(exam_result, data=request.data,)

#     try:
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#     except serializers.ValidationError as e:
#         return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
#     except Exception as e:
#         return Response(
#             {"detail": f"An error occurred: {str(e)}"},
#             status=status.HTTP_500_INTERNAL_SERVER_ERROR
#         )

#     return Response(serializer.data, status=status.HTTP_200_OK)


# @api_view(['PUT'])
# @permission_classes([CanChangeExamResult])
# def edit_exam_skill_result(request, exam_skill_result_id):
#     """
#     Update an ExamSkillResult (full update only).
#     """
#     try:
#         skill_result = ExamSkillResult.objects.get(exam_skill_result_id=exam_skill_result_id)
#     except ExamSkillResult.DoesNotExist:
#         return Response({"exam_skill_result_id": "Invalid ID"}, status=status.HTTP_404_NOT_FOUND)

#     serializer = EditExamSkillResultSerializer(skill_result, data=request.data,)

#     try:
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#     except serializers.ValidationError as e:
#         return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#     return Response(serializer.data, status=status.HTTP_200_OK)


# @api_view(['PATCH', 'PUT'])
# @permission_classes([CanChangeExamResult])
# def edit_exam_results(request, exam_result_id):
#     """
#     Update an ExamResult instance.
#     PATCH → Partial update
#     PUT   → Full update
#     Handles external/internal marks, co-scholastic grade, 
#     and automatically sets attendance based on marks.
#     """
#     try:
#         exam_result = ExamResult.objects.get(exam_result_id=exam_result_id, is_active=True)
#     except ExamResult.DoesNotExist:
#         return Response(
#             {"exam_result_id": "Invalid Exam Result ID"},
#             status=status.HTTP_404_NOT_FOUND
#         )

#     # If PUT, force full update (partial=False); PATCH → partial=True
#     partial_update = request.method == 'PATCH'

#     serializer = EditExamResultSerializer(exam_result, data=request.data, partial=partial_update)

#     try:
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#     except serializers.ValidationError as e:
#         return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
#     except Exception as e:
#         return Response(
#             {"detail": f"An error occurred: {str(e)}"},
#             status=status.HTTP_500_INTERNAL_SERVER_ERROR
#         )

#     return Response(serializer.data, status=status.HTTP_200_OK)

# @api_view(['PATCH', 'PUT'])
# @permission_classes([CanChangeExamResult])
# def edit_exam_skill_result(request, exam_skill_result_id):
#     """
#     Update an ExamSkillResult.
#     PATCH → partial update, PUT → full update
#     """
#     try:
#         skill_result = ExamSkillResult.objects.get(exam_skill_result_id=exam_skill_result_id)
#     except ExamSkillResult.DoesNotExist:
#         return Response({"exam_skill_result_id": "Invalid ID"}, status=status.HTTP_404_NOT_FOUND)

#     partial_update = request.method == 'PATCH'
#     serializer = EditExamSkillResultSerializer(skill_result, data=request.data, partial=partial_update)

#     try:
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#     except serializers.ValidationError as e:
#         return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
#     except Exception as e:
#         return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#     return Response(serializer.data, status=status.HTTP_200_OK)


class EditExamResultsViewSet(ModelViewSet):
    """
    ViewSet to retrieve and fully update a single ExamResult instance.
    Supports only GET and PUT.
    """
    queryset = ExamResult.objects.filter(is_active=True)
    permission_classes = [CanChangeExamResult]
    serializer_class = EditExamResultSerializer
    http_method_names = ['put']


        
class EditExamSkillResultViewSet(ModelViewSet):
    """
    ViewSet to retrieve and fully update a single ExamSkillResult instance.
    Supports only GET and PUT.
    """
    queryset = ExamSkillResult.objects.all()
    permission_classes = [CanChangeExamResult]
    serializer_class = EditExamSkillResultSerializer
    http_method_names = ['put']


class CoScholasticGradeDropdownViewSet(ModelViewSet):
    queryset = CoScholasticGrade.objects.filter(is_active=True).order_by('id')
    permission_classes = [IsAuthenticated]
    serializer_class = CoScholasticGradeDropdownSerializer
    http_method_names = ['get']

class CoScholasticGradeDropdownForMarksEntryViewSet(ModelViewSet):
    """
    Provides a dropdown list of Co-Scholastic Grades filtered by the exam's category.
    Requires `exam_id` query parameter.
    Example: /co_scholastic_grade_dropdown_for_marks_entry/?exam_id=<exam_id>
    """
    permission_classes = [IsAuthenticated]
    serializer_class = CoScholasticGradeDropdownSerializer
    http_method_names = ['get']

    def get_queryset(self):
        exam_id = self.request.query_params.get('exam_id')  # ✅ corrected method name
        if not exam_id:
            raise ParseError({'exam_id': "The 'exam_id' query parameter is required."})

        exam = get_object_or_404(Exam, exam_id=exam_id, is_active=True)

        return CoScholasticGrade.objects.filter(
            is_active=True,
            category=exam.category
        ).order_by('id')


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_marks_entry_expiry_datetime_in_exam_instance(request, exam_id):
    try:
        exam = Exam.objects.get(exam_id=exam_id, is_active=True)
    except Exam.DoesNotExist:
        return Response(
            {'exam_id': 'Invalid Exam ID'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ Take from JSON payload
    marks_entry_expiry_datetime_str = request.data.get('marks_entry_expiry_datetime')
    if not marks_entry_expiry_datetime_str:
        return Response(
            {'marks_entry_expiry_datetime': 'This field is required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ Convert string → datetime
    try:
        marks_entry_expiry_datetime = datetime.datetime.fromisoformat(marks_entry_expiry_datetime_str)
    except ValueError:
        return Response(
            {'marks_entry_expiry_datetime': 'Invalid datetime format. Use ISO format: YYYY-MM-DDTHH:MM:SS'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ Make timezone-aware if it's naive
    if timezone.is_naive(marks_entry_expiry_datetime):
        marks_entry_expiry_datetime = timezone.make_aware(marks_entry_expiry_datetime)

    now = timezone.now()
    if marks_entry_expiry_datetime <= now:
        return Response(
            {"marks_entry_expiry_datetime": "Marks entry expiry datetime must be in the future."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ Ensure expiry is after exam end date
    if exam.end_date:
        end_datetime = timezone.make_aware(datetime.datetime.combine(exam.end_date, datetime.time.min))
        if marks_entry_expiry_datetime <= end_datetime:
            return Response(
                {"marks_entry_expiry_datetime": "Marks entry expiry datetime must be after the exam end date."},
                status=status.HTTP_400_BAD_REQUEST
            )

    # ✅ Save updated datetime
    exam.marks_entry_expiry_datetime = marks_entry_expiry_datetime
    exam.save(update_fields=['marks_entry_expiry_datetime'])

    return Response(
        {'message': 'Marks Entry Expiry Date Updated Successfully'},
        status=status.HTTP_200_OK
    )


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def update_marks_entry_expiry_datetime_in_exam_instance(request, exam_id):
#     try:
#         exam = Exam.objects.get(exam_id=exam_id, is_active=True)
#     except Exam.DoesNotExist:
#         return Response({'exam_id': 'Invalid Exam ID'}, status=status.HTTP_400_BAD_REQUEST)

#     marks_entry_expiry_datetime_str = request.query_params.get('marks_entry_expiry_datetime')
#     if not marks_entry_expiry_datetime_str:
#         return Response({'marks_entry_expiry_datetime': 'marks_entry_expiry_datetime field is required'}, status=status.HTTP_400_BAD_REQUEST)

#     # ✅ Convert string → datetime
#     try:
#         marks_entry_expiry_datetime = datetime.datetime.fromisoformat(marks_entry_expiry_datetime_str)
#     except ValueError:
#         return Response(
#             {'marks_entry_expiry_datetime': 'Invalid datetime format. Use ISO format: YYYY-MM-DDTHH:MM:SS'},
#             status=status.HTTP_400_BAD_REQUEST
#         )

#     # ✅ Make timezone-aware if it's naive
#     if timezone.is_naive(marks_entry_expiry_datetime):
#         marks_entry_expiry_datetime = timezone.make_aware(marks_entry_expiry_datetime)

#     now = timezone.now()
#     if marks_entry_expiry_datetime <= now:
#         raise serializers.ValidationError({
#             "marks_entry_expiry_datetime": "Marks entry expiry datetime must be in the future."
#         })

#     # ✅ Ensure expiry is after exam end date (exact date, not end of day)
#     if exam.end_date:
#         end_datetime = timezone.make_aware(datetime.datetime.combine(exam.end_date, datetime.time.min))
#         if marks_entry_expiry_datetime <= end_datetime:
#             raise serializers.ValidationError({
#                 "marks_entry_expiry_datetime": "Marks entry expiry datetime must be after the exam end date."
#             })

#     exam.marks_entry_expiry_datetime = marks_entry_expiry_datetime
    
    
#     exam.save(update_fields=['marks_entry_expiry_datetime'])

#     return Response({'message': 'Marks Entry Expiry Date Updated Successfully'}, status=status.HTTP_200_OK)




#=========================================================================================================================================================
#==========================================    create exam instance     ==================================================================================
#=========================================================================================================================================================

# from rest_framework.decorators import api_view, permission_classes
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.response import Response
# from rest_framework import status
# from .models import ExamInstance
# from .serializers import ExamInstanceSerializer

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_exam_instance(request):
 
#     serializer = CreateExamInstanceSerializer(data=request.data)
#     if serializer.is_valid():
#         exam_id = serializer.validated_data.get('exam_id')
#         subject_id = serializer.validated_data.get('subject_id')

#         if not exam_id and  not subject_id:
#             return Response({"error": "Exam and subjects are requiered."},status=status.HTTP_400_BAD_REQUEST)


#         # ✅ Check for duplicate combination (exam + subject)
#         if ExamInstance.objects.filter(exam_id=exam_id, subject_id=subject_id).exists():
#             return Response({"error": "An ExamInstance with this Exam and Subject already exists."},status=status.HTTP_400_BAD_REQUEST)

#         # ✅ Save and assign created_by
#         serializer.save(subject_id = subject_id, created_by=request.user,updated_by = request.user)
#         return Response(serializer.data, status=status.HTTP_201_CREATED)

#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_exam_instance(request):
#     serializer = CreateExamInstanceSerializer(data=request.data)
#     if serializer.is_valid():
#         # serializer.validate already ensures exam is editable and uniqueness
#         serializer.save(created_by=request.user, updated_by=request.user)
#         return Response(serializer.data, status=status.HTTP_201_CREATED)
#     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_exam_instance(request, exam_id):  # take exam_id from URL
    data = request.data.copy()       # make a mutable copy of request data
    data['exam'] = exam_id           # inject exam_id into serializer data

    serializer = CreateExamInstanceSerializer(data=data)
    if serializer.is_valid():
        serializer.save(created_by=request.user, updated_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_exam_instance(request, pk):
    instance = get_object_or_404(ExamInstance, exam_instance_id = pk)
    # block update if exam is not editable (quick guard)
    if instance.exam and not getattr(instance.exam, "is_editable", True):
        return Response({"error": f"'{instance.exam.name}' is locked. You cannot update this ExamInstance."},
                        status=status.HTTP_400_BAD_REQUEST)

    serializer = CreateExamInstanceSerializer(instance, data=request.data, partial=False)
    if serializer.is_valid():
        serializer.save(updated_by=request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def marks_entry_expired_datetime_status(request):
    section_status_id = request.query_params.get('section_wise_exam_result_status_id')
    if not section_status_id:
        return Response({'section_wise_exam_result_status_id': "This field is required in the URL."},
                        status=status.HTTP_400_BAD_REQUEST)
    
    try:
        section_status = SectionWiseExamResultStatus.objects.select_related('exam', 'section').get(
            id=section_status_id, is_active=True
        )
    except SectionWiseExamResultStatus.DoesNotExist:
        return Response({'section_wise_exam_result_status_id': "Invalid id"},
                        status=status.HTTP_400_BAD_REQUEST)
    
    expiry_datetime = section_status.marks_entry_expiry_datetime

    # Convert to human-readable format
    human_readable = expiry_datetime.strftime('%Y-%m-%d %H:%M:%S') if expiry_datetime else None

    return Response({
        'marks_entry_expiry_datetime': expiry_datetime,
        'marks_entry_expiry_datetime_human': human_readable
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def marks_entry_percentage_for_marks_entry_page(request):
    section_status_id = request.query_params.get('section_wise_exam_result_status_id')
    if not section_status_id:
        return Response({'section_wise_exam_result_status_id': "This field is required in the URL."},
                        status=status.HTTP_400_BAD_REQUEST)
    
    try:
        section_status = SectionWiseExamResultStatus.objects.select_related('exam', 'section').get(
            id=section_status_id, is_active=True
        )
    except SectionWiseExamResultStatus.DoesNotExist:
        return Response({'section_wise_exam_result_status_id': "Invalid id"},
                        status=status.HTTP_400_BAD_REQUEST)
    
    marks_completion_percentage=section_status.marks_completion_percentage

    return Response({
        'marks_completion_percentage':marks_completion_percentage
    })

class ExamStatusDropDownViewset(ModelViewSet):
    queryset = ExamStatus.objects.filter(is_active=True).order_by('id')
    permission_classes = [IsAuthenticated]
    serializer_class = ExamStatusDropDropDownSerializer
    http_method_names = ['get']


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finalize_section_results(request):
    section_status_id = request.query_params.get('section_wise_exam_result_status_id')
    if not section_status_id:
        return Response(
            {'section_wise_exam_result_status_id': "This field is required in the URL."},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        section_status = SectionWiseExamResultStatus.objects.get(
            id=section_status_id, is_active=True
        )
    except SectionWiseExamResultStatus.DoesNotExist:
        return Response(
            {'section_wise_exam_result_status_id': "Invalid id"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if section_status.marks_completion_percentage != 100:
        return Response({
            'Section Status': f'Marks Entry Not Completed ({section_status.marks_completion_percentage}%)'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        finalized_status = ExamResultStatus.objects.get(id=4)
    except ExamResultStatus.DoesNotExist:
        return Response(
            {'status': "ExamResultStatus with id=3 not found."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Finalize section result
    section_status.finalized_by = request.user
    section_status.finalized_at = timezone.now()
    section_status.status = finalized_status
    section_status.save(update_fields=['finalized_by', 'finalized_at', 'status'])

    create_update_student_exam_summary.delay(section_status_id)

    return Response({
        'message': f'Section "{section_status.section.name}" results finalized successfully.'
    }, status=status.HTTP_200_OK)

class ExportBranchWiseExamResultStatusCSVViewSet(APIView):
    authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    search_fields = [
        'academic_year__name',
        'branch__name',
        'exam__name',
        'status__name',
        'exam__exam_type__name',
        'is_progress_card_downloaded',
    ]

    filterset_fields = [
        'academic_year',
        'branch',
        'exam',
        'status',
        'is_visible',
        'is_active',
        'exam__exam_type',
        'is_progress_card_downloaded',
    ]

    ordering_fields = [
        'academic_year__name',
        'branch__name',
        'exam__name',
        'status__name',
        'marks_entry_expiry_datetime',
        'marks_completion_percentage',
        'updated_at',
        'exam__exam_type__name',
        'is_progress_card_downloaded',
    ]
    ordering = ['-academic_year', 'marks_completion_percentage']

    filename = "Branch Wise Exam Results Marks Entry Status.csv"
    chunk_size = 1000

    def get(self, request, *args, **kwargs):
        user = self.request.user

        # ✅ Efficient branching logic
        if user.groups.filter(id=1).exists():  # Super admin or system user
            branches = Branch.objects.filter(is_active=True)
        else:
            branches = (
                UserProfile.objects.filter(user=user)
                .values_list('branches', flat=True)
                .distinct()
            )

        current_academic_year = AcademicYear.objects.filter(is_current_academic_year=True).first()
        if not current_academic_year:
            raise NotFound("Current academic year not found.")

        # ✅ Avoid returning inactive or invalid records
        queryset = (
            BranchWiseExamResultStatus.objects.filter(
                is_active=True,
                branch__in=branches,
                academic_year=current_academic_year,
            )
            .select_related('academic_year', 'branch', 'exam', 'status')  # optimization
            .order_by('-updated_at')
        )

        for backend in self.filter_backends:
            queryset = backend().filter_queryset(request, queryset, self)

        response = StreamingHttpResponse(
            self.generate_csv(queryset), content_type="text/csv"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}"'
        return response
    
    def generate_csv(self, queryset):
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        header = [ 'Sl.No.',
            "Academic Year", "Branch", "Exam Type", "Exam", "Status", "Marks Completion Percentage", 
            "Total Sections", "Pending Sections", "Completed Sections", "Marks Entry Expiry Date ",
        ]

        writer.writerow(header)
        buffer.seek(0)
        yield buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)

        sl_no = 1
        total = queryset.count()
        chunk_size = self.chunk_size

        # for obj in queryset.iterator(chunk_size=self.chunk_size):
        for start in range(0, total, chunk_size):
            chunk = queryset[start:start + chunk_size]
            for obj in chunk:
                academic_year = getattr(obj.academic_year, 'name', 'N/A') if obj.academic_year else 'N/A'
                branch = getattr(obj.branch, 'name' ,'N/A') if obj.branch else "N/A"
                exam_type = getattr(obj.exam.exam_type, 'name', 'N/A') if obj.exam.exam_type else 'N/A'
                exam = getattr(obj.exam, 'name', 'N/A') if obj.exam else 'N/A'
                status = getattr(obj.status, 'name')if obj.status else 'N/A'
                marks_entry_expiry_datetime = timezone.localtime(obj.marks_entry_expiry_datetime).strftime("%Y-%m-%d %H:%M:%S") if obj.marks_entry_expiry_datetime else ""

                row = [
                    sl_no,
                    academic_year,
                    branch,
                    exam_type,
                    exam,
                    status,
                    obj.marks_completion_percentage,
                    obj.total_sections,
                    obj.number_of_sections_pending,
                    obj.number_of_sections_completed,
                    marks_entry_expiry_datetime,
                ]

                writer.writerow(row)
                buffer.seek(0)
                yield buffer.getvalue()
                buffer.seek(0)
                buffer.truncate(0)
                sl_no += 1



# ========================================= Working upto removing rows and columns ====================================

import io
import csv
from django.http import StreamingHttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

class ExportSectionExamResultsCSVViewSet(APIView):
    authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filename_template = "{class_name} _{section} Section_Exam_Results.csv"
    chunk_size = 500

    def get(self, request, *args, **kwargs):
        section_status_id = request.query_params.get('section_wise_exam_result_status_id')
        if not section_status_id:
            return Response(
                {'section_wise_exam_result_status_id': "This field is required in the URL."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # === Fetch section & exam info ===
        section_status = (
            SectionWiseExamResultStatus.objects
            .select_related('exam', 'section__class_name')
            .filter(id=section_status_id, is_active=True)
            .first()
        )
        if not section_status:
            return Response({'section_wise_exam_result_status_id': "Invalid ID"}, status=400)

        exam = section_status.exam
        section = section_status.section

        # === Prefetch once ===
        exam_instances = list(
            ExamInstance.objects.filter(exam=exam, is_active=True).order_by('sequence')
            .prefetch_related('subject_skills')
        )
        skill_ids = exam_instances and list(
            ExamSubjectSkillInstance.objects.filter(exam_instance__in=exam_instances, is_active=True)
            .values_list('subject_skill_id', flat=True)
        )

        # === Map skill instances for quick access ===
        skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
            exam_instance__in=exam_instances, is_active=True, subject_skill_id__in=skill_ids
        ).select_related('subject_skill', 'exam_instance')
        skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

        students = list(
            Student.objects.filter(
                section=section,
                is_active=True,
                # academic_year=exam.academic_year,
            ).exclude(admission_status__admission_status_id=3)
        )

        student_ids = [s.student_id for s in students]
        exam_results = list(
            ExamResult.objects.filter(
                student_id__in=student_ids,
                exam_instance__in=exam_instances,
                is_active=True,
            ).select_related('exam_attendance', 'co_scholastic_grade')
        )
        exam_result_map = {(er.student_id, er.exam_instance_id): er for er in exam_results}

        exam_result_ids = [er.exam_result_id for er in exam_results]
        skill_results = list(
            ExamSkillResult.objects.filter(exam_result_id__in=exam_result_ids)
            .select_related('skill', 'exam_attendance', 'co_scholastic_grade')
        )
        skill_result_map = {(sr.exam_result_id, sr.skill_id): sr for sr in skill_results}

        # === Stream response ===
        filename = self.filename_template.format(
            class_name=section.class_name.name.replace(" ", "_"),
            section=section.name.replace(" ", "_"),
        )

        response = StreamingHttpResponse(
            self.generate_csv(students, exam_instances, skill_instance_map, exam_result_map, skill_result_map),
            content_type="text/csv"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def generate_csv(self, students, exam_instances, skill_instance_map, exam_result_map, skill_result_map):
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # === Dynamic header ===
        headers = ["Sl.No.", "Student Name", "SCS Number", "Marks Type"]
        external_row = internal_row = grade_row = False

        for instance in exam_instances:
            # Subject-level columns
            if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
                headers.append(instance.subject.name)
                if instance.has_external_marks:
                    external_row = True
                if instance.has_internal_marks:
                    internal_row = True
                if instance.has_subject_co_scholastic_grade:
                    grade_row = True
            if instance.has_subject_skills:
                # Skill-level columns
                for skill in instance.subject_skills.all():
                    si = skill_instance_map.get((instance.exam_instance_id, skill.id))
                    if si and (si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade):
                        headers.append(f"{instance.subject.name} - {skill.name}")
                        if si.has_external_marks:
                            external_row = True
                        if si.has_internal_marks:
                            internal_row = True
                        if si.has_subject_co_scholastic_grade:
                            grade_row = True

        writer.writerow(headers)
        yield from self._flush_buffer(buffer)

        # === Student rows ===
        for sl_no, student in enumerate(students, start=1):
            marks = {
                'external_marks': [''] * (len(headers) - 4),
                'internal_marks': [''] * (len(headers) - 4),
                'grade': [''] * (len(headers) - 4),
            }

            col_index = 0
            for instance in exam_instances:
                exam_result = exam_result_map.get((student.student_id, instance.exam_instance_id))
                if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
                    if exam_result:
                        att = exam_result.exam_attendance
                        if instance.has_external_marks:
                            marks['external_marks'][col_index] = (
                                exam_result.external_marks if att and att.exam_attendance_status_id == 1
                                else (att.short_code if att else '')
                            )
                        if instance.has_internal_marks:
                            marks['internal_marks'][col_index] = exam_result.internal_marks or ''
                        if instance.has_subject_co_scholastic_grade and exam_result.co_scholastic_grade:
                            marks['grade'][col_index] = exam_result.co_scholastic_grade.name
                    col_index += 1
                if instance.has_subject_skills:
                    for skill in instance.subject_skills.all():
                        si = skill_instance_map.get((instance.exam_instance_id, skill.id))
                        if not si or not exam_result:
                            col_index += 1
                            continue

                        sr = skill_result_map.get((exam_result.exam_result_id, skill.id))
                        if sr:
                            att = sr.exam_attendance
                            if si.has_external_marks:
                                marks['external_marks'][col_index] = (
                                    sr.external_marks if att and att.exam_attendance_status_id == 1
                                    else (att.short_code if att else '')
                                )
                            if si.has_internal_marks:
                                marks['internal_marks'][col_index] = sr.internal_marks or ''
                            if si.has_subject_co_scholastic_grade and sr.co_scholastic_grade:
                                marks['grade'][col_index] = sr.co_scholastic_grade.name
                        col_index += 1

            # === Write student data rows ===
            if external_row:
                writer.writerow([sl_no, student.name, student.SCS_Number, "External Marks"] + marks['external_marks'])
                yield from self._flush_buffer(buffer)
            if internal_row:
                writer.writerow([sl_no, student.name, student.SCS_Number, "Internal Marks"] + marks['internal_marks'])
                yield from self._flush_buffer(buffer)
            if grade_row:
                writer.writerow([sl_no, student.name, student.SCS_Number, "Grade"] + marks['grade'])
                yield from self._flush_buffer(buffer)

    def _flush_buffer(self, buffer):
        buffer.seek(0)
        data = buffer.read()
        yield data
        buffer.seek(0)
        buffer.truncate(0)

# import io
# import csv
# from django.http import StreamingHttpResponse
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework.permissions import IsAuthenticated
# from rest_framework import status
# from django_filters.rest_framework import DjangoFilterBackend
# from rest_framework.filters import SearchFilter, OrderingFilter

# class ExportSectionExamResultsCSVViewSet(APIView):
#     authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

#     filename_template = "{class_name} _{section} Section_Exam_Results.csv"
#     chunk_size = 500

#     def get(self, request, *args, **kwargs):
#         section_status_id = request.query_params.get('section_wise_exam_result_status_id')
#         if not section_status_id:
#             return Response(
#                 {'section_wise_exam_result_status_id': "This field is required in the URL."},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # === Fetch section & exam info ===
#         section_status = (
#             SectionWiseExamResultStatus.objects
#             .select_related('exam', 'section__class_name')
#             .filter(id=section_status_id, is_active=True)
#             .first()
#         )
#         if not section_status:
#             return Response({'section_wise_exam_result_status_id': "Invalid ID"}, status=400)

#         exam = section_status.exam
#         section = section_status.section

#         # === Prefetch once ===
#         exam_instances = list(
#             ExamInstance.objects.filter(exam=exam, is_active=True)
#             .prefetch_related('subject_skills')
#         )

#         skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
#             exam_instance__in=exam_instances, is_active=True
#         ).select_related('subject_skill', 'exam_instance')
#         skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

#         students = list(
#             Student.objects.filter(
#                 section=section,
#                 is_active=True,
#                 academic_year=exam.academic_year,
#             ).exclude(admission_status__admission_status_id=3)
#         )

#         student_ids = [s.student_id for s in students]
#         exam_results = list(
#             ExamResult.objects.filter(
#                 student_id__in=student_ids,
#                 exam_instance__in=exam_instances,
#                 is_active=True,
#             ).select_related('exam_attendance', 'co_scholastic_grade')
#         )
#         exam_result_map = {(er.student_id, er.exam_instance_id): er for er in exam_results}

#         exam_result_ids = [er.exam_result_id for er in exam_results]
#         skill_results = list(
#             ExamSkillResult.objects.filter(exam_result_id__in=exam_result_ids)
#             .select_related('skill', 'exam_attendance', 'co_scholastic_grade')
#         )
#         skill_result_map = {(sr.exam_result_id, sr.skill_id): sr for sr in skill_results}

#         # === Stream response ===
#         filename = self.filename_template.format(
#             class_name=section.class_name.name.replace(" ", "_"),
#             section=section.name.replace(" ", "_"),
#         )

#         response = StreamingHttpResponse(
#             self.generate_csv(students, exam_instances, skill_instance_map, exam_result_map, skill_result_map),
#             content_type="text/csv"
#         )
#         response["Content-Disposition"] = f'attachment; filename="{filename}"'
#         return response

#     def generate_csv(self, students, exam_instances, skill_instance_map, exam_result_map, skill_result_map):
#         buffer = io.StringIO()
#         writer = csv.writer(buffer)

#         # === Dynamic header ===
#         headers = ["Sl.No.", "Student Name", "SCS Number", "Marks Type"]
#         external_row = internal_row = grade_row = False

#         for instance in exam_instances:
#             # Subject-level columns
#             if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                 headers.append(instance.subject.name)
#                 if instance.has_external_marks:
#                     external_row = True
#                 if instance.has_internal_marks:
#                     internal_row = True
#                 if instance.has_subject_co_scholastic_grade:
#                     grade_row = True

#             # Skill-level columns
#             for skill in instance.subject_skills.all():
#                 si = skill_instance_map.get((instance.exam_instance_id, skill.id))
#                 if si and (si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade):
#                     headers.append(f"{instance.subject.name} - {skill.name}")
#                     if si.has_external_marks:
#                         external_row = True
#                     if si.has_internal_marks:
#                         internal_row = True
#                     if si.has_subject_co_scholastic_grade:
#                         grade_row = True

#         writer.writerow(headers)
#         yield from self._flush_buffer(buffer)

#         # === Student rows ===
#         for sl_no, student in enumerate(students, start=1):
#             marks = {
#                 'external_marks': [''] * (len(headers) - 4),
#                 'internal_marks': [''] * (len(headers) - 4),
#                 'grade': [''] * (len(headers) - 4),
#             }

#             col_index = 0
#             for instance in exam_instances:
#                 exam_result = exam_result_map.get((student.student_id, instance.exam_instance_id))
#                 if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                     if exam_result:
#                         att = exam_result.exam_attendance
#                         if instance.has_external_marks:
#                             marks['external_marks'][col_index] = (
#                                 exam_result.external_marks if att and att.exam_attendance_status_id == 1
#                                 else (att.short_code if att else '')
#                             )
#                         if instance.has_internal_marks:
#                             marks['internal_marks'][col_index] = exam_result.internal_marks or ''
#                         if instance.has_subject_co_scholastic_grade and exam_result.co_scholastic_grade:
#                             marks['grade'][col_index] = exam_result.co_scholastic_grade.name
#                     col_index += 1

#                 for skill in instance.subject_skills.all():
#                     si = skill_instance_map.get((instance.exam_instance_id, skill.id))
#                     if not si or not exam_result:
#                         col_index += 1
#                         continue

#                     sr = skill_result_map.get((exam_result.exam_result_id, skill.id))
#                     if sr:
#                         att = sr.exam_attendance
#                         if si.has_external_marks:
#                             marks['external_marks'][col_index] = (
#                                 sr.external_marks if att and att.exam_attendance_status_id == 1
#                                 else (att.short_code if att else '')
#                             )
#                         if si.has_internal_marks:
#                             marks['internal_marks'][col_index] = sr.internal_marks or ''
#                         if si.has_subject_co_scholastic_grade and sr.co_scholastic_grade:
#                             marks['grade'][col_index] = sr.co_scholastic_grade.name
#                     col_index += 1

#             # === Write student data rows ===
#             if external_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "External Marks"] + marks['external_marks'])
#                 yield from self._flush_buffer(buffer)
#             if internal_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "Internal Marks"] + marks['internal_marks'])
#                 yield from self._flush_buffer(buffer)
#             if grade_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "Grade"] + marks['grade'])
#                 yield from self._flush_buffer(buffer)

#     def _flush_buffer(self, buffer):
#         buffer.seek(0)
#         data = buffer.read()
#         yield data
#         buffer.seek(0)
#         buffer.truncate(0)

# class ExportSectionExamResultsCSVViewSet(APIView):
#     authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

#     filename = "Section_Exam_Results.csv"
#     chunk_size = 500

#     def get(self, request, *args, **kwargs):
#         section_status_id = request.query_params.get('section_wise_exam_result_status_id')
#         if not section_status_id:
#             return Response(
#                 {'section_wise_exam_result_status_id': "This field is required in the URL."},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         try:
#             section_status = SectionWiseExamResultStatus.objects.select_related('exam', 'section').get(
#                 id=section_status_id, is_active=True
#             )
#         except SectionWiseExamResultStatus.DoesNotExist:
#             return Response({'section_wise_exam_result_status_id': "Invalid id"},
#                             status=status.HTTP_400_BAD_REQUEST)

#         exam = section_status.exam
#         exam_instances = ExamInstance.objects.filter(exam=exam, is_active=True)
#         students = Student.objects.filter(
#             section=section_status.section,
#             is_active=True,
#             academic_year=exam.academic_year,
#         ).exclude(admission_status__admission_status_id=3)

#         response = StreamingHttpResponse(
#             self.generate_csv(students, exam_instances),
#             content_type="text/csv"
#         )
#         response["Content-Disposition"] = f'attachment; filename="{self.filename}".csv'
#         return response

#     def generate_csv(self, students, exam_instances):
#         buffer = io.StringIO()
#         writer = csv.writer(buffer)

#         # ===== Dynamic Header =====
#         dynamic_headers = []
#         external_row = internal_row = grade_row = False

#         for instance in exam_instances:
#             if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                 dynamic_headers.append(instance.subject.name)
#                 if instance.has_external_marks:
#                     external_row = True
#                 if instance.has_internal_marks:
#                     internal_row = True
#                 if instance.has_subject_co_scholastic_grade:
#                     grade_row = True

#             if instance.has_subject_skills:
#                 for skill in instance.subject_skills.all():
#                     skill_instance = ExamSubjectSkillInstance.objects.filter(
#                         exam_instance=instance, subject_skill=skill, is_active=True
#                     ).first()
#                     if (skill_instance and (
#                         skill_instance.has_external_marks or
#                         skill_instance.has_internal_marks or
#                         skill_instance.has_subject_co_scholastic_grade
#                     )):
#                         dynamic_headers.append(f'{instance.subject.name} - {skill.name}')
#                         if skill_instance.has_external_marks:
#                             external_row = True
#                         if skill_instance.has_internal_marks:
#                             internal_row = True
#                         if skill_instance.has_subject_co_scholastic_grade:
#                             grade_row = True

#         header = ['Sl.No.', 'Student Name', 'SCS Number', 'Marks Type'] + dynamic_headers
#         writer.writerow(header)
#         yield from self._flush_buffer(buffer, writer)

#         # ===== Write Data =====
#         for sl_no, student in enumerate(students, start=1):
#             # Prebuild empty lists to maintain column structure
#             marks = {
#                 'external_marks': [''] * len(dynamic_headers),
#                 'internal_marks': [''] * len(dynamic_headers),
#                 'grade': [''] * len(dynamic_headers),
#             }

#             column_index = 0
#             for instance in exam_instances:
#                 exam_result = ExamResult.objects.filter(
#                     student=student, exam_instance=instance, is_active=True
#                 ).select_related('co_scholastic_grade', 'exam_attendance').first()

#                 if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                     if exam_result:
#                         attendance = exam_result.exam_attendance
#                         # External
#                         if instance.has_external_marks:
#                             marks['external_marks'][column_index] = (
#                                 exam_result.external_marks if attendance.exam_attendance_status_id == 1
#                                 else attendance.short_code
#                             )
#                         # Internal
#                         if instance.has_internal_marks:
#                             marks['internal_marks'][column_index] = exam_result.internal_marks or ''
#                         # Grade
#                         if instance.has_subject_co_scholastic_grade and exam_result.co_scholastic_grade:
#                             marks['grade'][column_index] = exam_result.co_scholastic_grade.name
#                     column_index += 1

#                 # Skill-level marks
#                 if instance.has_subject_skills:
#                     for skill in instance.subject_skills.all():
#                         skill_instance = ExamSubjectSkillInstance.objects.filter(
#                             exam_instance=instance, subject_skill=skill, is_active=True
#                         ).first()
#                         if not skill_instance:
#                             continue

#                         skill_result = ExamSkillResult.objects.filter(
#                             exam_result=exam_result, skill=skill
#                         ).select_related('co_scholastic_grade', 'exam_attendance').first()

#                         if skill_result:
#                             attendance = skill_result.exam_attendance
#                             # External
#                             if skill_instance.has_external_marks:
#                                 marks['external_marks'][column_index] = (
#                                     skill_result.external_marks if attendance.exam_attendance_status_id == 1
#                                     else attendance.short_code
#                                 )
#                             # Internal
#                             if skill_instance.has_internal_marks:
#                                 marks['internal_marks'][column_index] = skill_result.internal_marks or ''
#                             # Grade
#                             if skill_instance.has_subject_co_scholastic_grade and skill_result.co_scholastic_grade:
#                                 marks['grade'][column_index] = skill_result.co_scholastic_grade.name
#                         column_index += 1

#             # ===== Write rows based on flags =====
#             if external_row:
#                 writer.writerow([
#                     sl_no, student.name, student.SCS_Number, "External Marks"
#                 ] + marks['external_marks'])
#                 yield from self._flush_buffer(buffer, writer)

#             if internal_row:
#                 writer.writerow([
#                     sl_no, student.name, student.SCS_Number, "Internal Marks"
#                 ] + marks['internal_marks'])
#                 yield from self._flush_buffer(buffer, writer)

#             if grade_row:
#                 writer.writerow([
#                     sl_no, student.name, student.SCS_Number, "Grade"
#                 ] + marks['grade'])
#                 yield from self._flush_buffer(buffer, writer)

#     def _flush_buffer(self, buffer, writer):
#         buffer.seek(0)
#         data = buffer.getvalue()
#         yield data
#         buffer.seek(0)
#         buffer.truncate(0)


#============================== Branch Exam Results Download ====================================

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from django.http import StreamingHttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

class BranchSectionsExamResultsXLSXView(APIView):
    authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]

    filename_template = "{branch}_{exam}_Exam_Results.xlsx"

    def get(self, request, *args, **kwargs):
        branch_wise_exam_result_status_id = request.query_params.get('branch_wise_exam_result_status_id')
        if not branch_wise_exam_result_status_id:
            return Response({'branch_wise_exam_result_status_id': "This field is required in the URL."}, status=400)

        branch_status = BranchWiseExamResultStatus.objects.select_related('branch', 'exam', 'academic_year').filter(
            id=branch_wise_exam_result_status_id,
            is_active=True
        ).first()
        if not branch_status:
            return Response({'branch_wise_exam_result_status_id': "Invalid Branch Wise Exam Result Status ID."}, status=400)

        exam = branch_status.exam

        # Fetch all sections belonging to this branch and exam
        sections = Section.objects.filter(
            academic_year=branch_status.academic_year,
            branch=branch_status.branch,
            class_name__class_name_id__in=exam.student_classes.values_list('class_name_id', flat=True),
            orientation__orientation_id__in=exam.orientations.values_list('orientation_id', flat=True),
            is_active=True,
            has_students=True
        ).distinct().order_by('class_name__class_sequence', 'name')

        if not sections.exists():
            return Response({
                "message": ["No sections found matching the given branch, exam classes, and orientations."]
            }, status=status.HTTP_404_NOT_FOUND)

        # Fetch exam instances
        exam_instances = list(ExamInstance.objects.filter(exam=exam, is_active=True).order_by('sequence').prefetch_related('subject_skills'))
        skill_ids = exam_instances and list(
                    ExamSubjectSkillInstance.objects.filter(exam_instance__in=exam_instances, is_active=True)
                    .values_list('subject_skill_id', flat=True)
                )

        # === Map skill instances for quick access ===
        skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
            exam_instance__in=exam_instances, is_active=True, subject_skill_id__in=skill_ids
        ).select_related('subject_skill', 'exam_instance')
        skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

        student_ids = Student.objects.filter(section__in=sections, is_active=True).values_list('student_id', flat=True)
        exam_results_qs = ExamResult.objects.filter(
            student_id__in=student_ids, exam_instance__in=exam_instances, is_active=True
        ).select_related('exam_attendance', 'co_scholastic_grade')
        exam_result_map = {(er.student_id, er.exam_instance_id): er for er in exam_results_qs}

        exam_result_ids = [er.exam_result_id for er in exam_results_qs]
        skill_results_qs = ExamSkillResult.objects.filter(
            exam_result_id__in=exam_result_ids
        ).select_related('exam_attendance', 'co_scholastic_grade', 'skill')
        skill_result_map = {(sr.exam_result_id, sr.skill_id): sr for sr in skill_results_qs}

        # ===== Workbook Setup =====
        wb = Workbook()
        wb.remove(wb.active)
        header_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for section in sections:
            ws_title = f"{section.class_name.name}_{section.name}"[:31]
            ws = wb.create_sheet(title=ws_title)

            # ===== Dynamic Header =====
            headers = ["Sl.No.", "Student Name", "SCS Number", "Marks Type"]
            external_row = internal_row = grade_row = False

            for instance in exam_instances:
                if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
                    headers.append(instance.subject.name)
                    if instance.has_external_marks:
                        external_row = True
                    if instance.has_internal_marks:
                        internal_row = True
                    if instance.has_subject_co_scholastic_grade:
                        grade_row = True
                if instance.has_subject_skills:
                    for skill in instance.subject_skills.all():
                        si = skill_instance_map.get((instance.exam_instance_id, skill.id))
                        if si and (si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade):
                            headers.append(f"{instance.subject.name} - {skill.name}")
                            if si.has_external_marks:
                                external_row = True
                            if si.has_internal_marks:
                                internal_row = True
                            if si.has_subject_co_scholastic_grade:
                                grade_row = True

            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = center

            # ===== Students and Results =====
            students = Student.objects.filter(
                section=section, is_active=True
            ).exclude(admission_status__admission_status_id=3)

            sl_no = 1
            for student in students:
                # Precompute mark values
                marks = {
                    'external_marks': [''] * (len(headers) - 4),
                    'internal_marks': [''] * (len(headers) - 4),
                    'grade': [''] * (len(headers) - 4),
                }
                col_index = 0

                for instance in exam_instances:
                    exam_result = exam_result_map.get((student.student_id, instance.exam_instance_id))
                    if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
                        if exam_result:
                            attendance = exam_result.exam_attendance
                            # External
                            if instance.has_external_marks:
                                marks['external_marks'][col_index] = (
                                    exam_result.external_marks if attendance and attendance.exam_attendance_status_id == 1
                                    else (attendance.short_code if attendance else '')
                                )
                            # Internal
                            if instance.has_internal_marks:
                                marks['internal_marks'][col_index] = exam_result.internal_marks or ''
                            # Grade
                            if instance.has_subject_co_scholastic_grade and exam_result.co_scholastic_grade:
                                marks['grade'][col_index] = exam_result.co_scholastic_grade.name
                        col_index += 1
                    if instance.has_subject_skills:
                        # Skills
                        for skill in instance.subject_skills.all():
                            si = skill_instance_map.get((instance.exam_instance_id, skill.id))
                            if not si:
                                continue
                            skill_result = exam_result and skill_result_map.get((exam_result.exam_result_id, skill.id))
                            if skill_result:
                                attendance = skill_result.exam_attendance
                                # External
                                if si.has_external_marks:
                                    marks['external_marks'][col_index] = (
                                        skill_result.external_marks if attendance and attendance.exam_attendance_status_id == 1
                                        else (attendance.short_code if attendance else '')
                                    )
                                # Internal
                                if si.has_internal_marks:
                                    marks['internal_marks'][col_index] = skill_result.internal_marks or ''
                                # Grade
                                if si.has_subject_co_scholastic_grade and skill_result.co_scholastic_grade:
                                    marks['grade'][col_index] = skill_result.co_scholastic_grade.name
                            col_index += 1

                # Rows to write
                row_start = ws.max_row + 1
                row_types = []
                if external_row:
                    row_types.append(("External Marks", marks['external_marks']))
                if internal_row:
                    row_types.append(("Internal Marks", marks['internal_marks']))
                if grade_row:
                    row_types.append(("Grade", marks['grade']))

                # Write student rows
                for i, (mark_type, mark_values) in enumerate(row_types):
                    ws.append([sl_no if i == 0 else "", student.name if i == 0 else "", student.SCS_Number if i == 0 else "", mark_type] + mark_values)

                # Merge cells vertically for Sl.No., Student Name, SCS Number
                if len(row_types) > 1:
                    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start + len(row_types) - 1, end_column=1)  # Sl.No.
                    ws.merge_cells(start_row=row_start, start_column=2, end_row=row_start + len(row_types) - 1, end_column=2)  # Student Name
                    ws.merge_cells(start_row=row_start, start_column=3, end_row=row_start + len(row_types) - 1, end_column=3)  # SCS Number

                sl_no += 1

        # ===== Stream XLSX =====
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = self.filename_template.format(
            branch=branch_status.branch.name.replace(" ", "_")[:20],
            exam=exam.name.replace(" ", "_")[:20],
        )

        response = StreamingHttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    
# import io
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font
# from django.http import StreamingHttpResponse
# from rest_framework.views import APIView
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.response import Response


# class BranchSectionsExamResultsXLSXView(APIView):
#     authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]

#     filename_template = "{branch}_{exam}_Exam_Results.xlsx"

#     def get(self, request, *args, **kwargs):
#         branch_wise_exam_result_status_id = request.query_params.get('branch_wise_exam_result_status_id')
#         if not branch_wise_exam_result_status_id:
#             return Response({'branch_wise_exam_result_status_id': "This field is required in the URL."}, status=400)

#         branch_status = BranchWiseExamResultStatus.objects.select_related('branch', 'exam', 'academic_year').filter(
#             id=branch_wise_exam_result_status_id,
#             is_active=True
#         ).first()
#         if not branch_status:
#             return Response({'branch_wise_exam_result_status_id': "Invalid Branch Wise Exam Result Status ID."}, status=400)

#         exam = branch_status.exam

#         # Fetch all sections belonging to this branch and exam
#         sections = Section.objects.filter(
#             academic_year=branch_status.academic_year,
#             branch=branch_status.branch,
#             class_name__class_name_id__in=exam.student_classes.values_list('class_name_id', flat=True),
#             orientation__orientation_id__in=exam.orientations.values_list('orientation_id', flat=True),
#             is_active=True,
#             has_students=True
#         ).distinct().order_by('class_name__class_sequence', 'name')

#         if not sections.exists():
#             return Response({
#                 "message": ["No sections found matching the given branch, exam classes, and orientations."]
#             }, status=status.HTTP_404_NOT_FOUND)

#         exam_instances = list(ExamInstance.objects.filter(exam=exam, is_active=True).prefetch_related('subject_skills'))

#         # Map skill and result data for quick lookup
#         skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
#             exam_instance__in=exam_instances, is_active=True
#         ).select_related('subject_skill', 'exam_instance')
#         skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

#         student_ids = Student.objects.filter(section__in=sections, is_active=True).values_list('student_id', flat=True)
#         exam_results_qs = ExamResult.objects.filter(
#             student_id__in=student_ids, exam_instance__in=exam_instances, is_active=True
#         ).select_related('exam_attendance', 'co_scholastic_grade')
#         exam_result_map = {(er.student_id, er.exam_instance_id): er for er in exam_results_qs}

#         exam_result_ids = [er.exam_result_id for er in exam_results_qs]
#         skill_results_qs = ExamSkillResult.objects.filter(
#             exam_result_id__in=exam_result_ids
#         ).select_related('exam_attendance', 'co_scholastic_grade', 'skill')
#         skill_result_map = {(sr.exam_result_id, sr.skill_id): sr for sr in skill_results_qs}

#         # ===== Workbook Setup =====
#         wb = Workbook()
#         wb.remove(wb.active)
#         header_font = Font(bold=True)
#         center = Alignment(horizontal='center', vertical='center', wrap_text=True)

#         for section in sections:
#             ws_title = f"{section.class_name.name}_{section.name}"[:31]
#             ws = wb.create_sheet(title=ws_title)

#             # ===== Dynamic Header =====
#             headers = ["SCS Number", "Student Name", "Marks Type"]
#             external_row = internal_row = grade_row = False

#             for instance in exam_instances:
#                 if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                     headers.append(instance.subject.name)
#                     if instance.has_external_marks:
#                         external_row = True
#                     if instance.has_internal_marks:
#                         internal_row = True
#                     if instance.has_subject_co_scholastic_grade:
#                         grade_row = True

#                 for skill in instance.subject_skills.all():
#                     si = skill_instance_map.get((instance.exam_instance_id, skill.id))
#                     if si and (
#                         si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade
#                     ):
#                         headers.append(f"{instance.subject.name} - {skill.name}")
#                         if si.has_external_marks:
#                             external_row = True
#                         if si.has_internal_marks:
#                             internal_row = True
#                         if si.has_subject_co_scholastic_grade:
#                             grade_row = True

#             ws.append(headers)
#             for col_idx in range(1, len(headers) + 1):
#                 cell = ws.cell(row=1, column=col_idx)
#                 cell.font = header_font
#                 cell.alignment = center

#             # ===== Students and Results =====
#             students = Student.objects.filter(
#                 section=section, is_active=True
#             ).exclude(admission_status__admission_status_id=3)

#             sl_no = 1
#             for student in students:
#                 # Precompute mark values
#                 marks = {
#                     'external_marks': [''] * (len(headers) - 3),
#                     'internal_marks': [''] * (len(headers) - 3),
#                     'grade': [''] * (len(headers) - 3),
#                 }
#                 col_index = 0

#                 for instance in exam_instances:
#                     exam_result = exam_result_map.get((student.student_id, instance.exam_instance_id))
#                     if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                         if exam_result:
#                             attendance = exam_result.exam_attendance
#                             # External
#                             if instance.has_external_marks:
#                                 marks['external_marks'][col_index] = (
#                                     exam_result.external_marks if attendance and attendance.exam_attendance_status_id == 1
#                                     else (attendance.short_code if attendance else '')
#                                 )
#                             # Internal
#                             if instance.has_internal_marks:
#                                 marks['internal_marks'][col_index] = exam_result.internal_marks or ''
#                             # Grade
#                             if instance.has_subject_co_scholastic_grade and exam_result.co_scholastic_grade:
#                                 marks['grade'][col_index] = exam_result.co_scholastic_grade.name
#                         col_index += 1

#                     # Skills
#                     for skill in instance.subject_skills.all():
#                         si = skill_instance_map.get((instance.exam_instance_id, skill.id))
#                         if not si:
#                             continue
#                         skill_result = exam_result and skill_result_map.get((exam_result.exam_result_id, skill.id))
#                         if skill_result:
#                             attendance = skill_result.exam_attendance
#                             # External
#                             if si.has_external_marks:
#                                 marks['external_marks'][col_index] = (
#                                     skill_result.external_marks if attendance and attendance.exam_attendance_status_id == 1
#                                     else (attendance.short_code if attendance else '')
#                                 )
#                             # Internal
#                             if si.has_internal_marks:
#                                 marks['internal_marks'][col_index] = skill_result.internal_marks or ''
#                             # Grade
#                             if si.has_subject_co_scholastic_grade and skill_result.co_scholastic_grade:
#                                 marks['grade'][col_index] = skill_result.co_scholastic_grade.name
#                         col_index += 1

#                 # ===== Only add rows that are actually needed =====
#                 if external_row:
#                     ws.append([student.SCS_Number, student.name, "External Marks"] + marks['external_marks'])
#                 if internal_row:
#                     ws.append([student.SCS_Number, student.name, "Internal Marks"] + marks['internal_marks'])
#                 if grade_row:
#                     ws.append([student.SCS_Number, student.name, "Grade"] + marks['grade'])
#                 sl_no += 1

#         # ===== Stream XLSX =====
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         filename = self.filename_template.format(
#             branch=branch_status.branch.name.replace(" ", "_")[:20],
#             exam=exam.name.replace(" ", "_")[:20],
#         )

#         response = StreamingHttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         return response

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def publish_progress_card_for_exam(request, exam_id):
    try:
        exam = Exam.objects.get(exam_id=exam_id, is_active=True)
    except Exam.DoesNotExist:
        return Response({"detail": "Exam not found."}, status=status.HTTP_404_NOT_FOUND)

    # Ensure exam is locked
    if exam.exam_status_id != 3:
        return Response({'message': 'Exam is Not Yet Locked'}, status=status.HTTP_400_BAD_REQUEST)

    # Check progress card mapping
    progresscard_mapping = ExamProgressCardMapping.objects.filter(exam=exam, is_active=True).first()
    if not progresscard_mapping:
        return Response({'message': 'Progress Card Template Not Yet Mapped'}, status=status.HTTP_400_BAD_REQUEST)
    
    exam.is_progress_card_visible = True 
    exam.save()

    # Filter sections whose results are published (status_id = 4)
    section_qs = SectionWiseExamResultStatus.objects.filter(
        exam=exam, is_active=True, status_id=4
    )

    section_ids = list(section_qs.values_list('section__section_id', flat=True))

    # If no sections are ready, notify
    if not section_ids:
        return Response({'message': 'No sections with published results found.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Update progress card download flag
    section_qs.update(is_progress_card_downloaded=True)

    # Update StudentExamSummary for those sections
    StudentExamSummary.objects.filter(
        exam=exam,
        student__section__section_id__in=section_ids
    ).update(is_progresscard=True)

    return Response({
        'message': 'Progress Card Published Successfully',
        'sections_updated': len(section_ids)
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def view_exam_details(request, exam_id):
    try:
        exam = Exam.objects.get(exam_id=exam_id, is_active=True)
    except Exam.DoesNotExist:
        return Response({'exam_id': 'Invalid Exam ID'}, status=status.HTTP_400_BAD_REQUEST)

    serializer = ViewExamSerializer(exam)
    return Response(serializer.data, status=status.HTTP_200_OK)



# import io
# from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
# from django.http import HttpResponse
# from rest_framework.views import APIView
# from rest_framework.permissions import IsAuthenticated
# from rest_framework.response import Response
# from django_filters.rest_framework import DjangoFilterBackend
# from rest_framework.filters import SearchFilter, OrderingFilter

class ExportSectionExamResultsTemplateXLSXView(APIView):
    authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filename_template = "{class_name}_{section}_Section_Exam_Results.xlsx"

    def get(self, request, *args, **kwargs):
        section_status_id = request.query_params.get("section_wise_exam_result_status_id")
        if not section_status_id:
            return Response(
                {"section_wise_exam_result_status_id": "This field is required in the URL."},
                status=400,
            )

        # === Fetch section & exam info ===
        section_status = (
            SectionWiseExamResultStatus.objects
            .select_related("exam", "section__class_name")
            .filter(id=section_status_id, is_active=True)
            .first()
        )
        if not section_status:
            return Response({"section_wise_exam_result_status_id": "Invalid ID"}, status=400)

        exam = section_status.exam
        section = section_status.section

        # === Prefetch Exam Instances ===
        exam_instances = list(
            ExamInstance.objects.filter(exam=exam, is_active=True).order_by('sequence')
            .prefetch_related("subject_skills")
        )
        skill_ids = exam_instances and list(
                    ExamSubjectSkillInstance.objects.filter(exam_instance__in=exam_instances, is_active=True)
                    .values_list('subject_skill_id', flat=True)
                )

        # === Map skill instances for quick access ===
        skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
            exam_instance__in=exam_instances, is_active=True, subject_skill_id__in=skill_ids
        ).select_related('subject_skill', 'exam_instance')
        skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

        students = list(
            Student.objects.filter(
                section=section,
                is_active=True,
                # academic_year=exam.academic_year,
            ).exclude(admission_status__admission_status_id=3)
        )

        # === Build Excel ===
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Results Template"

        # === Header Row ===
        headers = ["Sl.No.", "Student Name", "SCS Number", "Marks Type"]
        has_external = has_internal = has_grade = False

        for instance in exam_instances:
            if instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade:
                headers.append(instance.subject.name)
                has_external |= instance.has_external_marks
                has_internal |= instance.has_internal_marks
                has_grade |= instance.has_subject_co_scholastic_grade
            if instance.has_subject_skills:
                for skill in instance.subject_skills.all():
                    si = skill_instance_map.get((instance.exam_instance_id, skill.id))
                    if si and (
                        si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade
                    ):
                        headers.append(f"{instance.subject.name} - {skill.name}")
                        has_external |= si.has_external_marks
                        has_internal |= si.has_internal_marks
                        has_grade |= si.has_subject_co_scholastic_grade

        ws.append(headers)

        # === Style Header ===
        header_font = Font(bold=True)
        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        # === Fill Blank Template Rows (No Marks) ===
        row_num = 2
        for sl_no, student in enumerate(students, start=1):
            row_start = row_num
            marks_types = []

            if has_external:
                marks_types.append("External Marks")
            if has_internal:
                marks_types.append("Internal Marks")
            if has_grade:
                marks_types.append("Grade")

            for marks_type in marks_types:
                ws.cell(row=row_num, column=1, value=sl_no)
                ws.cell(row=row_num, column=2, value=student.name)
                ws.cell(row=row_num, column=3, value=student.SCS_Number)
                ws.cell(row=row_num, column=4, value=marks_type)
                row_num += 1

            # === Merge Student Info Cells Vertically ===
            if len(marks_types) > 1:
                ws.merge_cells(start_row=row_start, start_column=1, end_row=row_num - 1, end_column=1)
                ws.merge_cells(start_row=row_start, start_column=2, end_row=row_num - 1, end_column=2)
                ws.merge_cells(start_row=row_start, start_column=3, end_row=row_num - 1, end_column=3)

                for col in range(1, 4):
                    ws.cell(row=row_start, column=col).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )

        # === Border Styling ===
        thin = Side(border_style="thin", color="000000")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # === Return Response ===
        wb.save(output)
        output.seek(0)
        filename = self.filename_template.format(
            class_name=section.class_name.name.replace(" ", "_"),
            section=section.name.replace(" ", "_"),
        )

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

# class ExportSectionExamResultsTemplateXLSXView(APIView):
#     authentication_classes = [QueryParameterTokenAuthentication, SessionAuthentication]
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

#     filename_template = "{class_name} _{section} Section_Exam_Results.csv"
#     chunk_size = 500

#     def get(self, request, *args, **kwargs):
#         section_status_id = request.query_params.get('section_wise_exam_result_status_id')
#         if not section_status_id:
#             return Response(
#                 {'section_wise_exam_result_status_id': "This field is required in the URL."},
#                 status=status.HTTP_400_BAD_REQUEST
#             )

#         # === Fetch section & exam info ===
#         section_status = (
#             SectionWiseExamResultStatus.objects
#             .select_related('exam', 'section__class_name')
#             .filter(id=section_status_id, is_active=True)
#             .first()
#         )
#         if not section_status:
#             return Response({'section_wise_exam_result_status_id': "Invalid ID"}, status=400)

#         exam = section_status.exam
#         section = section_status.section

#         # === Prefetch once ===
#         exam_instances = list(
#             ExamInstance.objects.filter(exam=exam, is_active=True)
#             .prefetch_related('subject_skills')
#         )

#         skill_instances_qs = ExamSubjectSkillInstance.objects.filter(
#             exam_instance__in=exam_instances, is_active=True
#         ).select_related('subject_skill', 'exam_instance')
#         skill_instance_map = {(si.exam_instance_id, si.subject_skill_id): si for si in skill_instances_qs}

#         students = list(
#             Student.objects.filter(
#                 section=section,
#                 is_active=True,
#                 academic_year=exam.academic_year,
#             ).exclude(admission_status__admission_status_id=3)
#         )

#         student_ids = [s.student_id for s in students]

#         # === Stream response ===
#         filename = self.filename_template.format(
#             class_name=section.class_name.name.replace(" ", "_"),
#             section=section.name.replace(" ", "_"),
#         )

#         response = StreamingHttpResponse(
#             self.generate_csv(students, exam_instances, skill_instance_map),
#             content_type="text/csv"
#         )
#         response["Content-Disposition"] = f'attachment; filename="{filename}"'
#         return response
    
#     def generate_csv(self, students, exam_instances, skill_instance_map):
#         buffer = io.StringIO()
#         writer = csv.writer(buffer)

#         # === Dynamic header ===
#         headers = ["Sl.No.", "Student Name", "SCS Number", "Marks Type"]
#         external_row = internal_row = grade_row = False

#         for instance in exam_instances:
#             # Subject-level columns
#             if (instance.has_external_marks or instance.has_internal_marks or instance.has_subject_co_scholastic_grade):
#                 headers.append(instance.subject.name)
#                 if instance.has_external_marks:
#                     external_row = True
#                 if instance.has_internal_marks:
#                     internal_row = True
#                 if instance.has_subject_co_scholastic_grade:
#                     grade_row = True

#             # Skill-level columns
#             for skill in instance.subject_skills.all():
#                 si = skill_instance_map.get((instance.exam_instance_id, skill.id))
#                 if si and (si.has_external_marks or si.has_internal_marks or si.has_subject_co_scholastic_grade):
#                     headers.append(f"{instance.subject.name} - {skill.name}")
#                     if si.has_external_marks:
#                         external_row = True
#                     if si.has_internal_marks:
#                         internal_row = True
#                     if si.has_subject_co_scholastic_grade:
#                         grade_row = True

#         writer.writerow(headers)
#         yield from self._flush_buffer(buffer)

#         # === Student rows ===
#         for sl_no, student in enumerate(students, start=1):
#             # === Write student data rows ===
#             if external_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "External Marks"] )
#                 yield from self._flush_buffer(buffer)
#             if internal_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "Internal Marks"] )
#                 yield from self._flush_buffer(buffer)
#             if grade_row:
#                 writer.writerow([sl_no, student.name, student.SCS_Number, "Grade"])
#                 yield from self._flush_buffer(buffer)

#     def _flush_buffer(self, buffer):
#         buffer.seek(0)
#         data = buffer.read()
#         yield data
#         buffer.seek(0)
#         buffer.truncate(0)